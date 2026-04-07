from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
import re
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

import pandas as pd
from openpyxl import Workbook, load_workbook

from asin_isbn_conversion import asin_isbn_conversion
from load_weekly_files import get_latest_sales_csv
import paths as process_paths

WEEKLY_TEMPLATE_HEADER_ROWS = 8
WEEKLY_EXPORT_COLUMNS = [
    "ASIN",
    "External ID",
    "Customer Orders",
    "Units Shipped",
    "Units at Amazon",
    "Open PO qty",
]
WEEKLY_TOTAL_COLUMNS = [
    "Customer Orders",
    "Units Shipped",
    "Units at Amazon",
    "Open PO qty",
]
WEEKLY_HEADER_ALIASES = {
    "ASIN": ["ASIN"],
    "External ID": ["External ID"],
    "Customer Orders": ["Customer Orders", "Customer orders"],
    "Units Shipped": ["Units Shipped", "Units shipped"],
    "Units at Amazon": ["Units at Amazon", "Units at amazon"],
    "Open PO qty": ["Open PO qty", "Open PO Qty"],
}


def parse_week_end_date_from_sales_file(sales_file: str | Path) -> datetime:
    sales_name = Path(sales_file).name
    match = re.search(r"_(\d{1,2}-\d{1,2}-\d{4})\.csv$", sales_name)
    if not match:
        raise ValueError(f"Could not parse week-ending date from sales file: {sales_name}")
    return datetime.strptime(match.group(1), "%m-%d-%Y")


def previous_weekly_report_template(target_path: Path) -> Path | None:
    candidates = sorted(
        [
            path
            for path in process_paths.AMAZON_WEEKLY_REPORTS_DIR.glob("w*.xlsx")
            if path.resolve() != target_path.resolve()
        ],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def copy_header_rows(template_ws, output_ws) -> None:
    for row in template_ws.iter_rows(min_row=1, max_row=WEEKLY_TEMPLATE_HEADER_ROWS):
        for cell in row:
            target = output_ws[cell.coordinate]
            target.value = cell.value
            if cell.has_style:
                target._style = copy(cell._style)
            if cell.number_format:
                target.number_format = cell.number_format
            if cell.font:
                target.font = copy(cell.font)
            if cell.fill:
                target.fill = copy(cell.fill)
            if cell.border:
                target.border = copy(cell.border)
            if cell.alignment:
                target.alignment = copy(cell.alignment)
            if cell.protection:
                target.protection = copy(cell.protection)

    for merged_range in template_ws.merged_cells.ranges:
        if merged_range.max_row <= WEEKLY_TEMPLATE_HEADER_ROWS:
            output_ws.merge_cells(str(merged_range))

    for col_letter, dim in template_ws.column_dimensions.items():
        output_ws.column_dimensions[col_letter].width = dim.width
        output_ws.column_dimensions[col_letter].hidden = dim.hidden

    for row_idx in range(1, WEEKLY_TEMPLATE_HEADER_ROWS + 1):
        output_ws.row_dimensions[row_idx].height = template_ws.row_dimensions[row_idx].height
        output_ws.row_dimensions[row_idx].hidden = template_ws.row_dimensions[row_idx].hidden

    output_ws.freeze_panes = template_ws.freeze_panes
    if template_ws.sheet_view.zoomScale:
        output_ws.sheet_view.zoomScale = template_ws.sheet_view.zoomScale


def weekly_header_map(output_ws) -> dict[str, int]:
    header_positions: dict[str, int] = {}
    for col_idx in range(1, output_ws.max_column + 1):
        value = output_ws.cell(row=WEEKLY_TEMPLATE_HEADER_ROWS, column=col_idx).value
        if value is None:
            continue
        header_positions[str(value).strip().casefold()] = col_idx
    return header_positions


def create_weekly_report(df: pd.DataFrame) -> Path:
    sales_file = get_latest_sales_csv()
    if not sales_file:
        raise FileNotFoundError("Could not find the latest Amazon sales CSV.")

    week_end_date = parse_week_end_date_from_sales_file(sales_file)
    output_path = process_paths.amazon_weekly_report_file(week_end_date)
    template_path = previous_weekly_report_template(output_path)
    if template_path is None:
        raise FileNotFoundError(
            f"No prior weekly workbook was found in {process_paths.AMAZON_WEEKLY_REPORTS_DIR}"
        )

    process_paths.AMAZON_WEEKLY_REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    template_wb = load_workbook(template_path)
    try:
        template_ws = template_wb.active

        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = template_ws.title

        copy_header_rows(template_ws, output_ws)
        header_positions = weekly_header_map(output_ws)

        weekly_df = df[WEEKLY_EXPORT_COLUMNS].copy()
        weekly_df = weekly_df.fillna("")

        for row_idx, row_values in enumerate(
            weekly_df.itertuples(index=False, name=None), start=WEEKLY_TEMPLATE_HEADER_ROWS + 1
        ):
            for field_name, value in zip(WEEKLY_EXPORT_COLUMNS, row_values):
                target_col = None
                for alias in WEEKLY_HEADER_ALIASES.get(field_name, [field_name]):
                    target_col = header_positions.get(alias.casefold())
                    if target_col is not None:
                        break
                if target_col is None:
                    raise ValueError(
                        f'Header "{field_name}" was not found in row {WEEKLY_TEMPLATE_HEADER_ROWS} '
                        f"of template workbook {template_path.name}"
                    )
                output_ws.cell(row=row_idx, column=target_col, value=value)

        output_wb.save(output_path)
        output_wb.close()
        return output_path
    finally:
        template_wb.close()


def numeric_totals(df: pd.DataFrame, columns: list[str] | None = None) -> dict[str, float]:
    if columns is None:
        numeric_df = df.select_dtypes(include="number")
    else:
        numeric_df = df[columns].apply(pd.to_numeric, errors="coerce")

    totals: dict[str, float] = {}
    for column in numeric_df.columns:
        totals[column] = float(numeric_df[column].fillna(0).sum())
    return totals


def isbn_row_count(df: pd.DataFrame) -> int:
    return int(len(df.index))


def print_totals(title: str, totals: dict[str, float], isbn_count: int | None = None) -> None:
    print()
    print(title)
    if isbn_count is not None:
        print(f"  ISBN Row Count: {isbn_count:,}")
    for column, total in totals.items():
        if float(total).is_integer():
            display = f"{int(total):,}"
        else:
            display = f"{total:,.2f}"
        print(f"  {column}: {display}")

def main():
    df = asin_isbn_conversion()

    float_cols = df.select_dtypes(include="float64").columns.tolist()
    df = df[["ASIN", "ISBN"] + float_cols]

    rename = {
        "ISBN": "External ID",
        "Ordered Units": "Customer Orders",
        "Shipped Units": "Units Shipped",
        "Sellable On Hand Units": "Units at Amazon",
        "Open Purchase Order Quantity": "Open PO qty",
    }
    df = df.rename(columns=rename)

    df = df.drop_duplicates(subset=["ASIN", "External ID"])

    drop_asin = "1743791895"
    df = df[df["ASIN"] != drop_asin]

    df = df.drop_duplicates(subset="External ID", keep="first")
    total_isbn_rows = isbn_row_count(df)

    main_output_totals = numeric_totals(df)
    weekly_report_path = create_weekly_report(df)
    weekly_output_totals = numeric_totals(df, WEEKLY_TOTAL_COLUMNS)
    print(f"Weekly report workbook: {weekly_report_path}")

    root = tk.Tk()
    root.withdraw()

    default_output = process_paths.amazon_sql_upload_output_file()
    output_dir = process_paths.AMAZON_SQL_UPLOAD_OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Default output workbook: {default_output}")

    with pd.ExcelWriter(default_output) as writer:
        df.to_excel(writer, index=False, startrow=3)
    print(f"Saved Excel file: {default_output}")
    print_totals("Main Output Numeric Totals", main_output_totals, total_isbn_rows)

    print_totals("Weekly 6-Column Totals", weekly_output_totals, total_isbn_rows)
    print()

    no_isbn_count = (df["External ID"] == "NO_ISBN").sum()
    cols_to_show = [
        "ASIN",
        "External ID",
        "Customer Orders",
        "Units Shipped",
        "Units at Amazon",
        "Open PO qty",
    ]
    df_no_isbn = df[df["External ID"] == "NO_ISBN"][cols_to_show]

    msg = (
        f"{no_isbn_count} titles have NO_ISBN.\n\n"
        "Would you like to:\n"
        "1. View them on screen\n"
        "2. Save them as Excel\n\n"
        "Enter 1 or 2:"
    )
    user_choice = simpledialog.askstring("NO_ISBN Titles", msg)

    if user_choice == "1":
        print(df_no_isbn.head(20))
    elif user_choice == "2":
        no_isbn_file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save NO_ISBN Titles Excel File",
            initialdir=str(output_dir),
        )
        if no_isbn_file:
            df_no_isbn.to_excel(no_isbn_file, index=False)
            messagebox.showinfo("Saved", f"Saved Excel file: {no_isbn_file}")
        else:
            print("Save cancelled.")
    else:
        print("No action taken.")

    update_dicts = simpledialog.askstring(
        "Update Dictionaries",
        "Would you like to update the ASIN removal list or ASIN-->ISBN manual key?\nType 'y' for yes or 'n' for no:",
    )
    if update_dicts and update_dicts.lower() == "y":
        import subprocess
        import sys

        script_dir = Path(__file__).resolve().parent
        subprocess.run(
            [sys.executable, str(script_dir / "asin_add_to_dictionaries.py")],
            check=True,
        )
    else:
        print("Skipping dictionary update.")


if __name__ == "__main__":
    main()
