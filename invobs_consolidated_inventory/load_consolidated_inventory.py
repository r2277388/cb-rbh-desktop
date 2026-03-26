import pandas as pd
import numpy as np
import sys
from pathlib import Path
from tkinter import Tk
import openpyxl
from tkinter.filedialog import askopenfilename

# Function to load the data
def load_data(file_path, sheet_name):
    df = pd.read_excel(file_path,
                       skiprows=3,
                       na_values=[''],
                       sheet_name=sheet_name,
                       keep_default_na=False)
    df = df.fillna(0)
    df.columns = [
        "Item", "pgrp", "bkft", "title", "price", "ship", "ans", "uc_cbc", "uc_hbg", 
        "uc_cbp", "uc_all", "val_cbc", "units_cbc", "val_hbg", "units_hbg", 
        "val_cbp", "units_cbp", "total_Units", "total_value"
    ]
    return df

# Function to apply filters
def apply_filters(df, pgrp_filter, bkft_filter):
    df = df[df['pgrp'].isin(pgrp_filter)]
    df = df[df['bkft'].isin(bkft_filter)]
    df.reset_index(drop=True, inplace=True)
    return df


def choose_data_sheet(sheet_names):
    preferred_prefix = "all_consolidated_inventories_v_"
    preferred_matches = [
        sheet_name
        for sheet_name in sheet_names
        if sheet_name.lower().startswith(preferred_prefix)
    ]
    if preferred_matches:
        return preferred_matches[0]

    for sheet_name in sheet_names:
        lower_name = sheet_name.lower()
        if "consolidated" in lower_name:
            return sheet_name

    return sheet_names[0] if sheet_names else None

# # Main function to execute the steps
# def consolidate_inventory():
#     print(">>> consolidate_inventory() started")
#     # Open a file dialog to select the input file
#     Tk().withdraw()  # Hide the root Tkinter window
#     print(">>> File dialog opening...")
#     file_consolidated_inventory = askopenfilename(
#         title="Select the Consolidated Inventory File",
#         filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
#     )
    
#     print(f">>> File selected: {file_consolidated_inventory}")
    
#     if not file_consolidated_inventory:
#         print("No file selected. Exiting.")
#         return None

#     print(">>> Asking for sheet name 1")
#     # Prompt the user for the sheet name
#     sheet_name = input("Enter the name of the sheet containing the data: ")
#     print(f">>> Sheet name received: {sheet_name}")

#     # Load data
#     df = load_data(file_consolidated_inventory, sheet_name)
    
#     # Filters
#     pgrp_filter = ['ART', 'LIF', 'CHL', 'ENT', 'FWN', 'PTC', 'RID', 'GAM', 'BAR-LIF', 'BAR-ENT', 'BAR-ART', 'CCB', 'CPB','CPA']
#     bkft_filter = ['BK', 'FT','CP','RP']
    
#     # Apply filters
#     df_filtered = apply_filters(df, pgrp_filter, bkft_filter)
    
#     # Rename column correctly
#     df_filtered = df_filtered.rename(columns={'Item': 'ISBN'})
    
#     # Ensure ISBN is 13 digits
#     df_filtered['ISBN'] = df_filtered['ISBN'].astype(str).str.zfill(13)
    
#     return df_filtered

def consolidate_inventory(file_consolidated_inventory=None):
    print(">>> consolidate_inventory() started")

    if file_consolidated_inventory is None:
        Tk().withdraw()
        print(">>> File dialog opening...")
        file_consolidated_inventory = askopenfilename(
            title="Select the Consolidated Inventory File",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        print(f">>> File selected: {file_consolidated_inventory}")
    else:
        file_consolidated_inventory = str(Path(file_consolidated_inventory))
        print(f">>> Using provided file: {file_consolidated_inventory}")

    if not file_consolidated_inventory:
        print("No file selected. Exiting.")
        return None

    # List available sheet names
    wb = openpyxl.load_workbook(file_consolidated_inventory, read_only=True)
    sheetnames = wb.sheetnames
    sheet_name = choose_data_sheet(sheetnames)
    if sheet_name is None:
        print("No worksheets were found. Exiting.")
        return None

    print(f">>> Sheet selected: {sheet_name}")

    # Load data
    print(f">>> Loading data from sheet: {sheet_name}")
    df = load_data(file_consolidated_inventory, sheet_name)
    print(">>> Data loaded!")
    print("Loaded DataFrame shape:", df.shape)

    # Filters
    pgrp_filter = ['ART', 'LIF', 'CHL', 'ENT', 'FWN', 'PTC', 'RID', 'GAM', 'BAR-LIF', 'BAR-ENT', 'BAR-ART', 'CCB', 'CPB','CPA']
    bkft_filter = ['BK', 'FT','CP','RP']

    # Apply filters
    print(">>> Applying filters...")
    df_filtered = apply_filters(df, pgrp_filter, bkft_filter)
    print(">>> Filters applied.")
    print("Filtered DataFrame shape:", df_filtered.shape)

    # Rename column correctly
    print(">>> Renaming column...")
    df_filtered = df_filtered.rename(columns={'Item': 'ISBN'})
    print(">>> Column renamed.")
    print("Filtered DataFrame shape after renaming:", df_filtered.shape)

    # Ensure ISBN is 13 digits
    print(">>> Padding ISBN...")
    df_filtered['ISBN'] = df_filtered['ISBN'].astype(str).str.zfill(13)
    print(">>> ISBN padded.")
    print("Filtered DataFrame shape after padding:", df_filtered.shape)
    print(df_filtered.info())

    return df_filtered

def main():
    print(">>> You should NOT see this if running from main3.py")
    df = consolidate_inventory()
    print(df.shape)
    # if df is not None:
    #     print(df.info())
    #     print(df.head())

if __name__ == "__main__":
    main()
