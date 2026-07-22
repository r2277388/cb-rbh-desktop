from __future__ import annotations

import argparse
import re
import sys
import warnings
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.append(str(REPO_ROOT))

from amazon_rolling_reports.functions import build_column_totals, save_to_excel
from shared.bookscan_calendar import bookscan_parts, bookscan_week
from shared.pg_grouping import apply_pg_grouping

warnings.filterwarnings(
    "ignore",
    message="Cannot parse header or footer.*",
    category=UserWarning,
    module=r"openpyxl\.worksheet\.header_footer",
)

MODULE_DIR = Path(__file__).resolve().parent
SAMPLE_WORKBOOK = MODULE_DIR / "Week 26 - 2026 Rolling AWBC (062726).xlsx"
SOURCE_FOLDER = Path(r"F:\ANALYSIS\Finance\DataWarehouse\Weekly reports\2026\AWBC")
OUTPUT_FOLDER = Path(r"G:\SALES\2026 Sales Reports\Sell-Through Reporting\AWBC")
LOCAL_OUTPUT_FOLDER = MODULE_DIR / "output"
CACHE_DIR = MODULE_DIR / "cache"
POS_CACHE_FILE = CACHE_DIR / "awbc_pos_cache.parquet"
METADATA_CACHE_FILE = CACHE_DIR / "awbc_metadata_cache.parquet"
YEARLY_CACHE_FILE = CACHE_DIR / "awbc_yearly_history.parquet"

HEADER_ROW_INDEX = 5
PRE_DATE_COLUMN_COUNT = 20
WEEKLY_OUTPUT_START = pd.Timestamp("2019-01-01")
SOURCE_FILE_RE = re.compile(r"(?P<mmddyy>\d{6})\s+AWBC\.xlsx$", re.IGNORECASE)
SOURCE_NUMERIC_FIELDS = [
    "Retail",
    "Week1Units",
    "Week2Units",
    "YTD_Units",
    "BAM_OnHand",
    "Warehouse_OnHand",
    "Qty_OnOrder",
]
REPORT_AUDIT_FIELDS = [
    "BAM OH",
    "WH OH",
    "OO Qty",
    "OH_Avg",
    "W52",
    "6Wk Avg",
    "TYTD",
    "LYTD",
    "YTD Var",
    "LYFY",
    "LTD",
]


@dataclass(frozen=True)
class BuildResult:
    output_file: Path
    latest_week: pd.Timestamp
    source_files_processed: list[Path]
    report_shape: tuple[int, int]
    cache_rows: int


def normalize_isbn_series(series: pd.Series) -> pd.Series:
    values = series.astype("string").str.strip()
    values = values.str.replace(r"\.0$", "", regex=True)
    values = values.str.replace(r"\D", "", regex=True)
    values = values.mask(values.eq(""))
    return values


def _numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0)


def _read_parquet_or_empty(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    return pd.read_parquet(path)


def _write_cache(df: pd.DataFrame, path: Path) -> None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    df.to_parquet(path, index=False)


def _clean_text_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    df = df.copy()
    for column in columns:
        if column in df.columns:
            df[column] = df[column].astype("string").fillna("")
    return df


def _parse_source_week(path: Path) -> pd.Timestamp | None:
    match = SOURCE_FILE_RE.search(path.name)
    if not match:
        return None
    parsed = pd.to_datetime(match.group("mmddyy"), format="%m%d%y", errors="coerce")
    if pd.isna(parsed):
        return None
    return pd.Timestamp(parsed).normalize()


def seed_caches_from_sample(sample_workbook: Path = SAMPLE_WORKBOOK) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if not sample_workbook.exists():
        raise FileNotFoundError(f"AWBC seed workbook not found: {sample_workbook}")

    print(f"Seeding AWBC caches from: {sample_workbook}")
    raw = pd.read_excel(sample_workbook, sheet_name=0, header=HEADER_ROW_INDEX, dtype=object)
    raw = raw.dropna(how="all").copy()
    raw["ISBN"] = normalize_isbn_series(raw["ISBN13"])
    raw = raw[raw["ISBN"].notna()].copy()

    metadata = pd.DataFrame(
        {
            "ISBN": raw["ISBN"],
            "Pub": raw.get("PUB"),
            "pt": raw.get("PT"),
            "ft": raw.get("CAT"),
            "pgrp": raw.get("PGR"),
            "Title": raw.get("TITLE"),
            "Price": _numeric(raw.get("PRICE", pd.Series(index=raw.index))),
            "PubDate": pd.to_datetime(raw.get("SHIP"), errors="coerce"),
            "BAM_OnHand": _numeric(raw.get("BAM O/H", pd.Series(index=raw.index))),
            "Warehouse_OnHand": _numeric(raw.get("WH O/H", pd.Series(index=raw.index))),
            "Qty_OnOrder": _numeric(raw.get("O/O", pd.Series(index=raw.index))),
            "OH_Avg": _numeric(raw.get("O/H Avg Wk", pd.Series(index=raw.index))),
        }
    )
    metadata = metadata.drop_duplicates(subset=["ISBN"], keep="last")
    metadata = _clean_text_columns(metadata, ["ISBN", "Pub", "pt", "ft", "pgrp", "Title"])

    week_columns: list[object] = []
    year_columns: list[object] = []
    for column in raw.columns[18:]:
        if str(column).isdigit() and 2008 <= int(column) <= 2015:
            year_columns.append(column)
            continue
        parsed = pd.to_datetime(column, errors="coerce")
        if not pd.isna(parsed):
            week_columns.append(column)
            continue

    sales_frames = []
    for column in week_columns:
        week = pd.Timestamp(pd.to_datetime(column)).normalize()
        values = pd.DataFrame({"ISBN": raw["ISBN"], "Week": week, "Qty": _numeric(raw[column])})
        values = values[values["Qty"].ne(0)]
        sales_frames.append(values)
    pos = pd.concat(sales_frames, ignore_index=True) if sales_frames else pd.DataFrame(columns=["ISBN", "Week", "Qty"])
    pos = pos.groupby(["ISBN", "Week"], as_index=False)["Qty"].sum()

    yearly_frames = []
    for column in year_columns:
        year = int(column)
        values = pd.DataFrame({"ISBN": raw["ISBN"], "Year": year, "Qty": _numeric(raw[column])})
        values = values[values["Qty"].ne(0)]
        yearly_frames.append(values)
    yearly = (
        pd.concat(yearly_frames, ignore_index=True)
        if yearly_frames
        else pd.DataFrame(columns=["ISBN", "Year", "Qty"])
    )
    if not yearly.empty:
        yearly = yearly.groupby(["ISBN", "Year"], as_index=False)["Qty"].sum()

    _write_cache(pos, POS_CACHE_FILE)
    _write_cache(metadata, METADATA_CACHE_FILE)
    _write_cache(yearly, YEARLY_CACHE_FILE)
    return pos, metadata, yearly


def _source_files() -> list[tuple[pd.Timestamp, Path]]:
    if not SOURCE_FOLDER.exists():
        return []
    files: list[tuple[pd.Timestamp, Path]] = []
    for path in SOURCE_FOLDER.glob("*AWBC.xlsx"):
        if path.name.startswith("~$"):
            continue
        week = _parse_source_week(path)
        if week is not None:
            files.append((week, path))
    return sorted(files, key=lambda item: item[0])


def read_weekly_source(path: Path, week: pd.Timestamp) -> tuple[pd.DataFrame, pd.DataFrame]:
    source = pd.read_excel(path, sheet_name=0, dtype=object)
    source["ISBN"] = normalize_isbn_series(source["ISBN"])
    source = source[source["ISBN"].notna()].copy()

    sales = pd.DataFrame({"ISBN": source["ISBN"], "Week": week, "Qty": _numeric(source["Week1Units"])})
    sales = sales.groupby(["ISBN", "Week"], as_index=False)["Qty"].sum()

    metadata = pd.DataFrame(
        {
            "ISBN": source["ISBN"],
            "Pub": source.get("Pub_Group"),
            "pt": pd.NA,
            "ft": pd.NA,
            "pgrp": pd.NA,
            "Title": source.get("Title"),
            "Price": _numeric(source.get("Retail", pd.Series(index=source.index))),
            "PubDate": pd.to_datetime(source.get("IDate"), errors="coerce"),
            "BAM_OnHand": _numeric(source.get("BAM_OnHand", pd.Series(index=source.index))),
            "Warehouse_OnHand": _numeric(source.get("Warehouse_OnHand", pd.Series(index=source.index))),
            "Qty_OnOrder": _numeric(source.get("Qty_OnOrder", pd.Series(index=source.index))),
            "OH_Avg": pd.NA,
        }
    )
    metadata = metadata.groupby("ISBN", as_index=False).agg(
        {
            "Pub": "first",
            "pt": "first",
            "ft": "first",
            "pgrp": "first",
            "Title": "first",
            "Price": "max",
            "PubDate": "first",
            "BAM_OnHand": "sum",
            "Warehouse_OnHand": "sum",
            "Qty_OnOrder": "sum",
            "OH_Avg": "first",
        }
    )
    return sales, metadata


def refresh_caches(full_refresh: bool = False) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list[Path]]:
    if full_refresh or not POS_CACHE_FILE.exists() or not METADATA_CACHE_FILE.exists() or not YEARLY_CACHE_FILE.exists():
        pos, metadata, yearly = seed_caches_from_sample()
    else:
        pos = _read_parquet_or_empty(POS_CACHE_FILE)
        metadata = _read_parquet_or_empty(METADATA_CACHE_FILE)
        yearly = _read_parquet_or_empty(YEARLY_CACHE_FILE)

    if not pos.empty:
        pos["Week"] = pd.to_datetime(pos["Week"]).dt.normalize()
    latest_cache_week = pd.to_datetime(pos["Week"]).max() if not pos.empty else None
    processed: list[Path] = []

    for week, path in _source_files():
        if latest_cache_week is not None and week <= latest_cache_week:
            continue
        if week.weekday() != 5:
            print(f"Skipping non-Saturday AWBC source file: {path.name}")
            continue
        print(f"Adding AWBC week {week.strftime('%Y-%m-%d')} from: {path.name}")
        source_sales, source_metadata = read_weekly_source(path, week)
        pos = pd.concat([pos, source_sales], ignore_index=True)
        pos = pos.drop_duplicates(subset=["ISBN", "Week"], keep="last")

        metadata = update_metadata(metadata, source_metadata)
        processed.append(path)
        latest_cache_week = week

    # On-hand and on-order are point-in-time snapshots. Clear values for ISBNs
    # absent from the newest file so old inventory cannot leak into this week.
    if latest_cache_week is not None:
        latest_path = next(
            (path for week, path in reversed(_source_files()) if week == latest_cache_week),
            None,
        )
        if latest_path is not None:
            _, latest_metadata = read_weekly_source(latest_path, latest_cache_week)
            for column in ["BAM_OnHand", "Warehouse_OnHand", "Qty_OnOrder"]:
                metadata[column] = 0
            metadata = update_metadata(metadata, latest_metadata)

    _write_cache(pos, POS_CACHE_FILE)
    _write_cache(metadata, METADATA_CACHE_FILE)
    _write_cache(yearly, YEARLY_CACHE_FILE)
    return pos, metadata, yearly, processed


def update_metadata(existing: pd.DataFrame, updates: pd.DataFrame) -> pd.DataFrame:
    if existing.empty:
        return updates.copy()
    existing = existing.copy().set_index("ISBN", drop=False)
    updates = updates.copy().set_index("ISBN", drop=False)
    for isbn, update in updates.iterrows():
        if isbn not in existing.index:
            existing.loc[isbn, updates.columns] = update
            continue
        for column in ["Price", "PubDate", "BAM_OnHand", "Warehouse_OnHand", "Qty_OnOrder"]:
            existing.at[isbn, column] = update[column]
        for column in ["Pub", "pt", "ft", "pgrp", "Title", "OH_Avg"]:
            current = existing.at[isbn, column] if column in existing.columns else pd.NA
            incoming = update[column]
            if pd.isna(current) or current == "":
                existing.at[isbn, column] = incoming
    return existing.reset_index(drop=True)


def _metric_pivot(pos: pd.DataFrame, mask: pd.Series, name: str, divisor: float = 1.0) -> pd.Series:
    values = pos.loc[mask].groupby("ISBN")["Qty"].sum()
    if divisor != 1.0:
        values = values / divisor
    return values.rename(name)


def build_report_dataframe(pos: pd.DataFrame, metadata: pd.DataFrame, yearly_history: pd.DataFrame) -> tuple[pd.DataFrame, pd.Timestamp]:
    if pos.empty:
        raise RuntimeError("AWBC POS cache is empty.")

    pos = pos.copy()
    pos["Week"] = pd.to_datetime(pos["Week"]).dt.normalize()
    pos["Qty"] = _numeric(pos["Qty"])
    latest_week = pd.Timestamp(pos["Week"].max()).normalize()
    latest_bookscan = bookscan_week(latest_week)

    parts = bookscan_parts(pos["Week"])
    pos = pd.concat([pos.reset_index(drop=True), parts.reset_index(drop=True)], axis=1)

    metadata = metadata.copy()
    metadata["ISBN"] = normalize_isbn_series(metadata["ISBN"])
    metadata = metadata[metadata["ISBN"].notna()].drop_duplicates(subset=["ISBN"], keep="last").set_index("ISBN")

    all_isbns = pd.Index(sorted(set(pos["ISBN"]) | set(metadata.index)), name="ISBN")
    metrics = pd.DataFrame(index=all_isbns)
    metrics["W52"] = _metric_pivot(pos, pos["Week"].ge(latest_week - pd.Timedelta(weeks=51)), "W52")
    metrics["6Wk Avg"] = _metric_pivot(pos, pos["Week"].ge(latest_week - pd.Timedelta(weeks=5)), "6Wk Avg", divisor=6.0)
    metrics["TYTD"] = _metric_pivot(
        pos,
        pos["BookScanYear"].eq(latest_bookscan.year) & pos["BookScanWeek"].le(latest_bookscan.week),
        "TYTD",
    )
    metrics["LYTD"] = _metric_pivot(
        pos,
        pos["BookScanYear"].eq(latest_bookscan.year - 1) & pos["BookScanWeek"].le(latest_bookscan.week),
        "LYTD",
    )
    metrics["LYFY"] = _metric_pivot(pos, pos["BookScanYear"].eq(latest_bookscan.year - 1), "LYFY")

    yearly_history = yearly_history.copy()
    if not yearly_history.empty:
        yearly_history["Year"] = pd.to_numeric(yearly_history["Year"], errors="coerce").astype("Int64")
        yearly_history["Qty"] = _numeric(yearly_history["Qty"])

    old_yearly_sum = (
        yearly_history.groupby("ISBN")["Qty"].sum()
        if not yearly_history.empty
        else pd.Series(dtype="float64")
    )
    metrics["LTD"] = pos.groupby("ISBN")["Qty"].sum().add(old_yearly_sum, fill_value=0)
    metrics["YTD Var"] = metrics["TYTD"].fillna(0) - metrics["LYTD"].fillna(0)
    metrics = metrics.fillna(0)

    weekly_pos = pos[pos["Week"].ge(WEEKLY_OUTPUT_START)].copy()
    weekly_pos["WeekLabel"] = weekly_pos["Week"].dt.strftime("%m-%d-%Y")
    weekly = weekly_pos.pivot_table(index="ISBN", columns="WeekLabel", values="Qty", aggfunc="sum", fill_value=0)
    week_order = sorted(weekly.columns, key=lambda value: pd.to_datetime(value, format="%m-%d-%Y"), reverse=True)
    weekly = weekly.reindex(columns=week_order)

    pre_2019 = pos[pos["Week"].dt.year.between(2016, 2018)].copy()
    pre_2019["Year"] = pre_2019["Week"].dt.year
    yearly_from_weeks = pre_2019.groupby(["ISBN", "Year"], as_index=False)["Qty"].sum()
    yearly_combined = pd.concat([yearly_from_weeks, yearly_history], ignore_index=True)
    yearly_combined = yearly_combined[yearly_combined["Year"].notna()].copy()
    yearly_combined["Year"] = yearly_combined["Year"].astype(int)
    yearly_wide = yearly_combined.pivot_table(index="ISBN", columns="Year", values="Qty", aggfunc="sum", fill_value=0)
    year_order = sorted([year for year in yearly_wide.columns if year <= 2018], reverse=True)
    yearly_wide = yearly_wide.reindex(columns=year_order)
    yearly_wide.columns = [str(column) for column in yearly_wide.columns]

    base_columns = ["Pub", "pt", "ft", "pgrp", "Title", "Price", "PubDate", "BAM_OnHand", "Warehouse_OnHand", "Qty_OnOrder", "OH_Avg"]
    for column in base_columns:
        if column not in metadata.columns:
            metadata[column] = pd.NA
    report = pd.concat([metadata.reindex(all_isbns)[base_columns], metrics, weekly.reindex(all_isbns), yearly_wide.reindex(all_isbns)], axis=1)
    report = report.reset_index()
    report = apply_pg_grouping(
        report,
        publisher_col="Pub",
        publishing_group_col="pgrp",
        product_type_col="pt",
        output_col="PG_Group",
    )
    report = report.rename(
        columns={
            "BAM_OnHand": "BAM OH",
            "Warehouse_OnHand": "WH OH",
            "Qty_OnOrder": "OO Qty",
        }
    )

    report["PubDate"] = pd.to_datetime(report["PubDate"], errors="coerce").dt.strftime("%m-%d-%Y").fillna("")
    for column in ["Pub", "pt", "ft", "pgrp", "PG_Group", "Title"]:
        report[column] = report[column].fillna("")
    for column in ["Price", "BAM OH", "WH OH", "OO Qty", "OH_Avg", "W52", "6Wk Avg", "TYTD", "LYTD", "YTD Var", "LYFY", "LTD"]:
        report[column] = _numeric(report[column])

    ordered = [
        "Pub",
        "pt",
        "ft",
        "pgrp",
        "PG_Group",
        "ISBN",
        "Title",
        "Price",
        "PubDate",
        "BAM OH",
        "WH OH",
        "OO Qty",
        "OH_Avg",
        "W52",
        "6Wk Avg",
        "TYTD",
        "LYTD",
        "YTD Var",
        "LYFY",
        "LTD",
    ]
    history_columns = [column for column in report.columns if column not in ordered]
    for column in history_columns:
        report[column] = _numeric(report[column])
    report = report[ordered + history_columns].copy()
    latest_label = latest_week.strftime("%m-%d-%Y")
    report = report.sort_values([latest_label, "Pub", "pgrp", "Title", "ISBN"], ascending=[False, True, True, True, True])
    return report.reset_index(drop=True), latest_week


def _format_output_filename(latest_week: pd.Timestamp) -> str:
    week = bookscan_week(latest_week).week
    return f"Week {week:02d} - {latest_week.year} Rolling AWBC ({latest_week.strftime('%m%d%y')}).xlsx"


def _history_columns(df: pd.DataFrame) -> list[str]:
    return [column for column in df.columns if isinstance(column, str) and (column.count("-") == 2 or column.isdigit())]


def _save_options(report_df: pd.DataFrame, latest_week: pd.Timestamp) -> dict[str, object]:
    history_columns = _history_columns(report_df)
    summary_columns = [
        "BAM OH",
        "WH OH",
        "OO Qty",
        "OH_Avg",
        "W52",
        "6Wk Avg",
        "TYTD",
        "LYTD",
        "YTD Var",
        "LYFY",
        "LTD",
    ] + history_columns
    decimal_columns = ["Price", "OH_Avg", "6Wk Avg"]
    accounting_format = {"num_format": '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'}
    decimal_format = {"num_format": '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'}
    column_format_overrides = {
        7: {"format": decimal_format, "width": 10},
        12: {"format": decimal_format, "width": 11},
        14: {"format": decimal_format, "width": 11},
    }
    for idx in range(9, min(PRE_DATE_COLUMN_COUNT, len(report_df.columns))):
        column_format_overrides.setdefault(idx, {"format": accounting_format})

    title_block = {
        "start_row": 1,
        "end_row": 2,
        "start_col": 6,
        "end_col": 6,
        "title": "AWBC Rolling POS",
        "subtitle": f"Week Ending: {latest_week.strftime('%B %d, %Y')}",
        "merge_cells": False,
        "align": "center",
    }
    return {
        "summary": build_column_totals(report_df, summary_columns),
        "format_cols": [column for column in summary_columns if column not in decimal_columns],
        "decimal_cols": decimal_columns,
        "integer_accounting_no_symbol": True,
        "rolling_main_layout": True,
        "pre_date_column_count": PRE_DATE_COLUMN_COUNT,
        "summary_label_col_idx": 8,
        "header_fill_overrides": {9: "#E6B8B7", 10: "#E6B8B7", 11: "#E6B8B7"},
        "format_blank_summary_cells": False,
        "title_block": title_block,
        "header_row_override": 5,
        "show_weeknum_label": True,
        "weeknum_label_fill": "#C4BD97",
        "column_format_overrides": column_format_overrides,
    }


def _source_numeric_totals(path: Path) -> dict[str, float]:
    source = pd.read_excel(path, sheet_name=0, usecols=SOURCE_NUMERIC_FIELDS)
    return {column: float(_numeric(source[column]).sum()) for column in SOURCE_NUMERIC_FIELDS}


def _report_total_row(path: Path) -> dict[object, float]:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore", message="Cannot parse header or footer.*", category=UserWarning
        )
        workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        top_rows = list(sheet.iter_rows(min_row=1, max_row=10, values_only=True))
        totals = next(
            (
                row
                for row in top_rows
                if any(value in {"Total", "Grand Total"} for value in row)
            ),
            top_rows[0],
        )
        headers = next(
            row
            for row in top_rows
            if any(value in {"BAM OH", "BAM O/H"} for value in row)
            and any(value in {"OO Qty", "O/O"} for value in row)
        )
        aliases = {
            "BAM O/H": "BAM OH",
            "WH O/H": "WH OH",
            "O/O": "OO Qty",
            "O/H Avg Wk": "OH_Avg",
            "52 WK": "W52",
            "YTD": "TYTD",
            "2025": "LYFY",
            2025: "LYFY",
        }
        return {aliases.get(header, header): value for header, value in zip(headers, totals)}
    finally:
        workbook.close()


def _weekly_total(total_row: dict[object, float], week: pd.Timestamp) -> float:
    for header, value in total_row.items():
        if isinstance(header, (pd.Timestamp,)) or hasattr(header, "year"):
            parsed = pd.Timestamp(header)
        elif isinstance(header, str):
            parsed = pd.to_datetime(header, format="%m-%d-%Y", errors="coerce")
        else:
            continue
        if not pd.isna(parsed) and pd.Timestamp(parsed).normalize() == week:
            return float(value or 0)
    return 0.0


def _total_numeric(value: object) -> float:
    parsed = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return 0.0 if pd.isna(parsed) else float(parsed)


def print_five_week_audit() -> None:
    recent_sources = list(reversed(_source_files()))[:5]
    if not recent_sources:
        return

    source_columns: dict[str, dict[str, float]] = {}
    report_columns: dict[str, dict[str, float]] = {}
    reconciliation: list[tuple[str, str]] = []
    direct_fields = {
        "Week1Units": "Weekly Sales",
        "BAM_OnHand": "BAM OH",
        "Warehouse_OnHand": "WH OH",
        "Qty_OnOrder": "OO Qty",
    }

    for week, source_path in recent_sources:
        label = week.strftime("%m/%d/%Y")
        source_totals = _source_numeric_totals(source_path)
        source_columns[label] = source_totals
        report_path = OUTPUT_FOLDER / _format_output_filename(week)
        if not report_path.exists():
            reconciliation.append((label, "REPORT MISSING"))
            continue

        total_row = _report_total_row(report_path)
        report_totals = {
            "Weekly Sales": _weekly_total(total_row, week),
            **{
                field: _total_numeric(total_row.get(field))
                for field in REPORT_AUDIT_FIELDS
            },
        }
        report_columns[label] = report_totals
        matches = all(
            abs(source_totals[source_field] - report_totals[report_field]) < 0.005
            for source_field, report_field in direct_fields.items()
        )
        reconciliation.append((label, "MATCH" if matches else "MISMATCH"))

    print("\nAWBC source totals - newest week first")
    print(pd.DataFrame(source_columns).to_string(float_format=lambda value: f"{value:,.2f}"))
    print("\nAWBC generated-report totals - newest week first")
    print(pd.DataFrame(report_columns).to_string(float_format=lambda value: f"{value:,.2f}"))
    print("\nAWBC direct-field reconciliation")
    for label, status in reconciliation:
        print(f"    {label}  {status}")


def build_awbc_report(full_refresh: bool = False, local_only: bool = False) -> BuildResult:
    pos, metadata, yearly, processed = refresh_caches(full_refresh=full_refresh)
    report, latest_week = build_report_dataframe(pos, metadata, yearly)
    output_folder = LOCAL_OUTPUT_FOLDER if local_only else OUTPUT_FOLDER
    output_folder.mkdir(parents=True, exist_ok=True)
    output_file = output_folder / _format_output_filename(latest_week)
    save_to_excel(report, output_file, **_save_options(report, latest_week))
    return BuildResult(
        output_file=output_file,
        latest_week=latest_week,
        source_files_processed=processed,
        report_shape=report.shape,
        cache_rows=len(pos),
    )


def print_status() -> None:
    pos = _read_parquet_or_empty(POS_CACHE_FILE)
    cache_week = pd.to_datetime(pos["Week"]).max() if not pos.empty and "Week" in pos.columns else None
    source_files = _source_files()
    latest_source = source_files[-1][0] if source_files else None
    print(f"Latest AWBC cache week: {cache_week.strftime('%Y-%m-%d') if cache_week is not None else 'None'}")
    print(f"Latest AWBC source file week: {latest_source.strftime('%Y-%m-%d') if latest_source is not None else 'None'}")
    if cache_week is not None and latest_source is not None and latest_source > cache_week:
        next_files = [path.name for week, path in source_files if week > cache_week]
        print("Pending source files:")
        for name in next_files:
            print(f"  - {name}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the AWBC weekly rolling POS report.")
    parser.add_argument("--full-refresh", action="store_true", help="Rebuild all AWBC caches from the seed workbook before adding source files.")
    parser.add_argument("--local-only", action="store_true", help="Write the output workbook under awbc_rolling_reports/output instead of the G: report folder.")
    parser.add_argument("--status", action="store_true", help="Show AWBC cache/source status without building the workbook.")
    args = parser.parse_args()

    if args.status:
        print_status()
        return

    print_status()
    result = build_awbc_report(full_refresh=args.full_refresh, local_only=args.local_only)
    print()
    print(f"AWBC report week: {result.latest_week.strftime('%Y-%m-%d')}")
    print(f"AWBC cache rows: {result.cache_rows:,}")
    if result.source_files_processed:
        print("Source files added to cache:")
        for path in result.source_files_processed:
            print(f"  - {path.name}")
    else:
        print("No newer AWBC source files needed to be added to cache.")
    print(f"Report shape: {result.report_shape[0]:,} rows x {result.report_shape[1]:,} columns")
    print(f"Saved report: {result.output_file}")
    print_five_week_audit()


if __name__ == "__main__":
    main()
