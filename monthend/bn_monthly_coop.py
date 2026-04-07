from __future__ import annotations

import argparse
from copy import copy
from datetime import datetime
from pathlib import Path
import re
import sys
import tkinter as tk
from tkinter import messagebox, simpledialog

import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

try:
    from .queries import barnes_noble_monthly_coop_sql
except ImportError:
    from queries import barnes_noble_monthly_coop_sql
from shared.db.connection import get_connection
from shared.db.query_runner import fetch_data_from_db


COLUMN_WIDTHS = {
    "A": 14.0,
    "B": 17.29,
    "C": 21.57,
    "D": 26.57,
    "E": 10.71,
    "F": 14.0,
    "G": 24.57,
    "H": 5.43,
    "I": 23.0,
    "J": 13.14,
    "K": 13.43,
    "L": 14.0,
    "M": 11.29,
    "N": 13.57,
    "O": 14.14,
    "P": 11.43,
    "Q": 17.0,
}
NUMERIC_COLUMNS = [
    "ISBN",
    "Retail Price",
    "Gross Units",
    "Return Units",
    "Net Units",
    "Gross Sales",
    "Return Sales",
    "Net Sales",
    "Net Retail Sales",
]
TEMPLATE_FILE = Path(__file__).resolve().parent / "202602_retail_monthly_coop_calc.xlsx"
ACCOUNTING_NO_SYMBOL_FORMAT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
TOTAL_FILL_HEX = "B7DEE8"
SUBTOTAL_FILL_HEX = "DAEEF3"


def prompt_for_period() -> str | None:
    root = tk.Tk()
    root.withdraw()
    try:
        while True:
            period = simpledialog.askstring(
                "Barnes & Noble Monthly Coop (Ailing)",
                "Enter the period in YYYYMM format:",
                parent=root,
            )
            if period is None:
                return None

            normalized = period.strip()
            if re.fullmatch(r"\d{6}", normalized):
                return normalized

            messagebox.showerror(
                "Invalid Period",
                "Please enter the period in YYYYMM format, for example 202603.",
                parent=root,
            )
    finally:
        root.destroy()


def output_path_for_period(period: str) -> Path:
    year = period[:4]
    output_dir = Path(fr"F:\{year}\Recurring Claims\B&N")
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"{period}_retail_monthly_coop_calc.xlsx"


def report_title(period: str) -> str:
    month_dt = datetime.strptime(period, "%Y%m")
    return f"B&N - Core Titles for {month_dt.strftime('%B %Y')}"


def fetch_monthly_coop_data(period: str) -> pd.DataFrame:
    engine = get_connection()
    query = barnes_noble_monthly_coop_sql(period)
    df = fetch_data_from_db(engine, query)

    if "Invoice Date" in df.columns:
        df["Invoice Date"] = pd.to_datetime(df["Invoice Date"])

    for column_name in NUMERIC_COLUMNS:
        if column_name in df.columns:
            df[column_name] = pd.to_numeric(df[column_name], errors="coerce")

    return df


def write_report(df: pd.DataFrame, period: str, output_path: Path) -> None:
    sql_query = barnes_noble_monthly_coop_sql(period)
    data_start_row = 4  # zero-based; Excel row 5
    header_row = 3      # zero-based; Excel row 4
    last_excel_row = len(df) + 4

    with pd.ExcelWriter(output_path, engine="xlsxwriter", datetime_format="mm/dd/yy") as writer:
        df.to_excel(writer, sheet_name="saldet", startrow=header_row, index=False)
        workbook = writer.book
        worksheet = writer.sheets["saldet"]
        sql_worksheet = workbook.add_worksheet("sql")
        writer.sheets["sql"] = sql_worksheet

        title_format = workbook.add_format({"bold": True, "font_size": 12})
        header_format = workbook.add_format(
            {
                "bold": True,
                "bg_color": "#D9EAD3",
                "border": 1,
            }
        )
        date_format = workbook.add_format({"num_format": "mm/dd/yy"})
        integer_format = workbook.add_format({"num_format": "#,##0"})
        money_format = workbook.add_format({"num_format": "#,##0.00"})
        total_label_format = workbook.add_format({"bold": True})

        worksheet.write("A1", report_title(period), title_format)
        worksheet.write("J1", "Totals", total_label_format)
        worksheet.write("J2", "Subtotals", total_label_format)

        for col_idx, column_name in enumerate(df.columns):
            worksheet.write(header_row, col_idx, column_name, header_format)

        if len(df) > 0:
            for col_letter in ["K", "L", "M", "N", "O", "P", "Q"]:
                worksheet.write_formula(
                    f"{col_letter}1",
                    f"=SUM({col_letter}5:{col_letter}{last_excel_row})",
                )
                worksheet.write_formula(
                    f"{col_letter}2",
                    f"=SUBTOTAL(9,{col_letter}5:{col_letter}{last_excel_row})",
                )
        else:
            for col_letter in ["K", "L", "M", "N", "O", "P", "Q"]:
                worksheet.write_blank(f"{col_letter}1", None)
                worksheet.write_blank(f"{col_letter}2", None)

        worksheet.freeze_panes(data_start_row, 0)
        worksheet.autofilter(header_row, 0, max(last_excel_row - 1, header_row), len(df.columns) - 1)

        for col_letter, width in COLUMN_WIDTHS.items():
            worksheet.set_column(f"{col_letter}:{col_letter}", width)

        if "Invoice Date" in df.columns:
            worksheet.set_column("A:A", COLUMN_WIDTHS["A"], date_format)

        column_positions = {name: idx for idx, name in enumerate(df.columns)}
        for column_name in ["Gross Units", "Return Units", "Net Units"]:
            if column_name in column_positions:
                idx = column_positions[column_name]
                worksheet.set_column(idx, idx, None, integer_format)

        for column_name in ["Retail Price", "Gross Sales", "Return Sales", "Net Sales", "Net Retail Sales"]:
            if column_name in column_positions:
                idx = column_positions[column_name]
                worksheet.set_column(idx, idx, None, money_format)

        sql_header_format = workbook.add_format({"bold": True})
        sql_body_format = workbook.add_format({"text_wrap": False})
        sql_worksheet.set_column("A:A", 5)
        sql_worksheet.set_column("B:B", 14)
        sql_worksheet.set_column("C:C", 140)

        sql_lines = sql_query.strip("\n").splitlines()
        for row_idx, line in enumerate(sql_lines, start=2):
            cleaned_line = line.rstrip()
            stripped = cleaned_line.strip()
            if not stripped:
                continue

            if not cleaned_line.startswith(" "):
                sql_worksheet.write_string(row_idx - 1, 1, stripped, sql_header_format)
            else:
                sql_worksheet.write_string(row_idx - 1, 2, stripped, sql_body_format)

    apply_example_styles(output_path, len(df.columns), last_excel_row)


def apply_example_styles(output_path: Path, max_columns: int, last_excel_row: int) -> None:
    if not TEMPLATE_FILE.exists():
        print(f"Template file not found, skipping style copy: {TEMPLATE_FILE}")
        return

    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    template_wb = load_workbook(TEMPLATE_FILE)
    template_ws = template_wb["saldet"]

    output_wb = load_workbook(output_path)
    output_ws = output_wb["saldet"]

    for source_cell, target_cell in [("A1", "A1"), ("J1", "J1"), ("J2", "J2")]:
        output_ws[target_cell]._style = copy(template_ws[source_cell]._style)
        output_ws[target_cell].font = copy(template_ws[source_cell].font)
        output_ws[target_cell].fill = copy(template_ws[source_cell].fill)
        output_ws[target_cell].border = copy(template_ws[source_cell].border)
        output_ws[target_cell].alignment = copy(template_ws[source_cell].alignment)
        output_ws[target_cell].protection = copy(template_ws[source_cell].protection)

    output_ws["B1"].fill = copy(template_ws["A1"].fill)

    from openpyxl.styles import PatternFill

    total_fill = PatternFill(fill_type="solid", fgColor=TOTAL_FILL_HEX)
    subtotal_fill = PatternFill(fill_type="solid", fgColor=SUBTOTAL_FILL_HEX)
    for target_cell in ["J1", "K1", "L1", "M1", "N1", "O1", "P1", "Q1"]:
        output_ws[target_cell].fill = copy(total_fill)
    for target_cell in ["J2", "K2", "L2", "M2", "N2", "O2", "P2", "Q2"]:
        output_ws[target_cell].fill = copy(subtotal_fill)

    for col_idx in range(1, max_columns + 1):
        col_letter = get_column_letter(col_idx)
        template_header = template_ws[f"{col_letter}4"]
        target_header = output_ws[f"{col_letter}4"]
        target_header._style = copy(template_header._style)
        target_header.font = copy(template_header.font)
        target_header.fill = copy(template_header.fill)
        target_header.border = copy(template_header.border)
        target_header.alignment = copy(template_header.alignment)
        target_header.protection = copy(template_header.protection)

        template_data = template_ws[f"{col_letter}5"]
        for row_idx in range(5, last_excel_row + 1):
            target_cell = output_ws[f"{col_letter}{row_idx}"]
            target_cell._style = copy(template_data._style)
            target_cell.font = copy(template_data.font)
            target_cell.fill = copy(template_data.fill)
            target_cell.border = copy(template_data.border)
            target_cell.alignment = copy(template_data.alignment)
            target_cell.protection = copy(template_data.protection)

        if template_ws.column_dimensions[col_letter].width is not None:
            output_ws.column_dimensions[col_letter].width = template_ws.column_dimensions[col_letter].width

    output_ws.freeze_panes = template_ws.freeze_panes
    output_ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines

    header_names = [output_ws.cell(row=4, column=col_idx).value for col_idx in range(1, max_columns + 1)]
    numeric_column_indexes = {
        index for index, name in enumerate(header_names, start=1) if name in NUMERIC_COLUMNS
    }

    for col_idx in numeric_column_indexes:
        col_letter = get_column_letter(col_idx)
        for row_idx in range(5, last_excel_row + 1):
            output_ws[f"{col_letter}{row_idx}"].number_format = ACCOUNTING_NO_SYMBOL_FORMAT
        for summary_row in [1, 2]:
            summary_cell = output_ws[f"{col_letter}{summary_row}"]
            if isinstance(summary_cell.value, str) and not summary_cell.value.startswith("="):
                continue
            summary_cell.number_format = ACCOUNTING_NO_SYMBOL_FORMAT

    isbn_column_indexes = {
        index for index, name in enumerate(header_names, start=1) if name == "ISBN"
    }
    for col_idx in isbn_column_indexes:
        col_letter = get_column_letter(col_idx)
        output_ws.column_dimensions[col_letter].width = 14.0
        for row_idx in range(5, last_excel_row + 1):
            output_ws[f"{col_letter}{row_idx}"].number_format = "General"

    output_wb.save(output_path)


def run(period: str | None = None) -> Path | None:
    selected_period = period or prompt_for_period()
    if not selected_period:
        print("No period selected. Exiting.")
        return None

    print(f"Running Barnes & Noble Monthly Coop (Ailing) for {selected_period}...")
    df = fetch_monthly_coop_data(selected_period)
    output_path = output_path_for_period(selected_period)
    write_report(df, selected_period, output_path)
    print(f"Saved report: {output_path}")
    return output_path


def main(argv: list[str] | None = None) -> Path | None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--period",
        help="Accounting period in YYYYMM format, for example 202603.",
    )
    args = parser.parse_args(argv)
    return run(args.period)


if __name__ == "__main__":
    main()
