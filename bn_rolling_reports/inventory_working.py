from __future__ import annotations

import argparse
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from config import (
    INVENTORY_DATA_START_ROW,
    INVENTORY_HEADER_ROW,
    INVENTORY_PREFIX,
    INVENTORY_REQUIRED_HEADERS,
    INVENTORY_TOTAL_ROW,
    format_legacy_bn_removed_isbns_filename,
    format_legacy_bn_output_filename,
)
from isbn_utils import load_bn_upload_isbns, normalize_isbn, normalize_isbn_series
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pos_combiner import format_output_filename, parse_week_ending, resolve_raw_folder

FOOTER_MARKERS = (
    "Data available from",
    "Copyright",
)


@dataclass
class InventoryBuildResult:
    raw_folder: Path
    source_file: Path
    output_file: Path
    removed_isbns_file: Path
    matched_updates: int
    appended_rows: int
    excluded_rows: int
    final_shape: tuple[int, int]


def format_inventory_output_filename(week_ending: datetime) -> str:
    return format_legacy_bn_output_filename(INVENTORY_PREFIX, week_ending)


def format_previous_inventory_output_filename(week_ending: datetime) -> str:
    return f"{INVENTORY_PREFIX}_{week_ending:%Y_%m_%d}.xlsx"


def find_inventory_source_file(raw_folder: Path) -> Path:
    matches = sorted(
        [
            path
            for path in raw_folder.iterdir()
            if path.is_file()
            and path.suffix.lower() == ".xlsx"
            and path.name.lower().startswith("inventory")
            and not path.name.startswith("~$")
        ],
        key=lambda path: path.name.lower(),
    )
    if not matches:
        raise FileNotFoundError(
            f"Could not find an Inventory*.xlsx file in {raw_folder}"
        )
    return matches[0]


def get_pos_output_file(raw_folder: Path) -> Path:
    week_ending = parse_week_ending(raw_folder.name)
    return raw_folder / format_output_filename(week_ending)


def load_pos_dataframe(raw_folder: Path) -> pd.DataFrame:
    pos_output = get_pos_output_file(raw_folder)
    if not pos_output.exists():
        raise FileNotFoundError(
            f"Combined POS file not found: {pos_output}. Run pos_combiner.py first."
        )

    df = pd.read_excel(pos_output, dtype={"ISBN": "string", "Imprint": "string"})
    required = {"ISBN", "Imprint", "OH_DC", "OH_Stores", "OO_DC", "OO_Stores"}
    missing = required.difference(df.columns)
    if missing:
        raise ValueError(
            f"Combined POS file is missing required columns: {sorted(missing)}"
        )

    df = df.loc[
        :, ["ISBN", "Imprint", "OH_DC", "OH_Stores", "OO_DC", "OO_Stores"]
    ].copy()
    df["ISBN"] = normalize_isbn_series(df["ISBN"])
    df = df.loc[df["ISBN"].notna()].drop_duplicates(subset=["ISBN"], keep="first")
    return df


def normalize_existing_headers(ws) -> None:
    current_headers = [
        ws.cell(row=INVENTORY_HEADER_ROW, column=idx).value for idx in range(1, 34)
    ]
    if current_headers == INVENTORY_REQUIRED_HEADERS:
        return

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= INVENTORY_HEADER_ROW and merged_range.max_row >= 1:
            ws.unmerge_cells(str(merged_range))

    ws.delete_cols(6, 6)

    for insert_at in (6, 10, 14, 18, 22, 26, 30):
        ws.insert_cols(insert_at, amount=1)

    for col_idx in range(6, 34):
        for row_idx in range(1, INVENTORY_HEADER_ROW + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

    row1_labels = {
        7: "DOH",
        9: "SOH",
        11: "DOO",
        13: "SOO",
    }
    row3_labels = {
        7: "Inventory",
        8: "Inventory",
        9: "Inventory",
        11: "On Order",
        12: "On Order",
        13: "On Order",
        15: "Due Out",
        16: "Due Out",
        17: "Due Out",
        19: "Reorder Threshold",
        20: "Reorder Threshold",
        21: "Reorder Threshold",
        23: "# of stores per weekly sales",
        24: "# of stores per weekly sales",
        25: "# of stores per weekly sales",
        27: "# of stores on hand",
        28: "# of stores on hand",
        29: "# of stores on hand",
        31: "# of stores per model",
        32: "# of stores per model",
        33: "# of stores per model",
    }
    for col_idx in (
        7, 8, 9, 11, 12, 13, 15, 16, 17, 19, 20, 21, 23, 24, 25, 27, 28, 29, 31, 32, 33
    ):
        ws.cell(row=4, column=col_idx).value = "Barnes & Noble"

    for col_idx, value in row1_labels.items():
        ws.cell(row=1, column=col_idx).value = value

    for col_idx, value in row3_labels.items():
        ws.cell(row=3, column=col_idx).value = value

    for col_idx, header in enumerate(INVENTORY_REQUIRED_HEADERS, start=1):
        ws.cell(row=INVENTORY_HEADER_ROW, column=col_idx).value = header


def remove_footer_rows(ws) -> None:
    for row_idx in range(INVENTORY_DATA_START_ROW, ws.max_row + 1):
        first_value = ws.cell(row=row_idx, column=1).value
        if isinstance(first_value, str) and any(
            first_value.strip().startswith(marker) for marker in FOOTER_MARKERS
        ):
            ws.delete_rows(row_idx, ws.max_row - row_idx + 1)
            return


def get_last_data_row(ws) -> int:
    for row_idx in range(ws.max_row, INVENTORY_DATA_START_ROW - 1, -1):
        row_has_data = any(
            ws.cell(row=row_idx, column=col_idx).value not in (None, "")
            for col_idx in range(1, 34)
        )
        if row_has_data:
            return row_idx
    return INVENTORY_DATA_START_ROW - 1


def clone_row_styles(ws, source_row: int, target_row: int, max_col: int = 33) -> None:
    for col_idx in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col_idx)
        target_cell = ws.cell(row=target_row, column=col_idx)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def build_existing_inventory_index(ws, last_data_row: int) -> dict[str, int]:
    index: dict[str, int] = {}
    for row_idx in range(INVENTORY_DATA_START_ROW, last_data_row + 1):
        raw_isbn = ws.cell(row=row_idx, column=1).value
        isbn = normalize_isbn(raw_isbn)
        if not isbn or isbn in index:
            continue
        index[isbn] = row_idx
        ws.cell(row=row_idx, column=1).value = isbn
        ws.cell(row=row_idx, column=1).number_format = "@"
    return index


def excel_safe_value(value):
    return None if pd.isna(value) else value


def append_missing_rows(
    ws, start_row: int, pos_missing: pd.DataFrame, style_row: int
) -> int:
    appended = 0
    for _, row in pos_missing.iterrows():
        target_row = start_row + appended
        clone_row_styles(ws, style_row, target_row)
        ws.cell(row=target_row, column=1).value = excel_safe_value(row["ISBN"])
        ws.cell(row=target_row, column=1).number_format = "@"
        ws.cell(row=target_row, column=3).value = excel_safe_value(row["Imprint"])
        ws.cell(row=target_row, column=7).value = excel_safe_value(row["OH_DC"])
        ws.cell(row=target_row, column=9).value = excel_safe_value(row["OH_Stores"])
        ws.cell(row=target_row, column=11).value = excel_safe_value(row["OO_DC"])
        ws.cell(row=target_row, column=13).value = excel_safe_value(row["OO_Stores"])
        appended += 1
    return appended


def remove_disallowed_rows(ws, last_data_row: int, allowed_isbns: set[str]) -> int:
    removed_rows: list[dict[str, object]] = []
    for row_idx in range(last_data_row, INVENTORY_DATA_START_ROW - 1, -1):
        isbn = normalize_isbn(ws.cell(row=row_idx, column=1).value)
        if not isbn or isbn not in allowed_isbns:
            removed_rows.append(
                {
                    "ISBN": isbn or "",
                    "Title": ws.cell(row=row_idx, column=2).value,
                    "Author": ws.cell(row=row_idx, column=4).value,
                    "Imprint": ws.cell(row=row_idx, column=3).value,
                    "Reason": "Missing/invalid ISBN" if not isbn else "Not in BN upload whitelist",
                }
            )
            ws.delete_rows(row_idx, 1)
    removed_rows.reverse()
    return removed_rows


def save_removed_isbns_report(removed_rows: list[dict[str, object]], output_file: Path) -> None:
    df = pd.DataFrame(
        removed_rows,
        columns=["ISBN", "Title", "Author", "Imprint", "Reason"],
    )
    output_file.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_file, index=False)


def refresh_grand_total_row(ws, last_data_row: int) -> None:
    target_columns = (7, 9, 11, 13)
    if last_data_row < INVENTORY_DATA_START_ROW:
        for col_idx in target_columns:
            ws.cell(row=INVENTORY_TOTAL_ROW, column=col_idx).value = 0
        return

    for col_idx in target_columns:
        col_letter = get_column_letter(col_idx)
        ws.cell(
            row=INVENTORY_TOTAL_ROW, column=col_idx
        ).value = (
            f"=SUM({col_letter}{INVENTORY_DATA_START_ROW}:{col_letter}{last_data_row})"
        )


def build_inventory_working_file(
    raw_folder: str | Path | None = None,
    output_file: str | Path | None = None,
) -> InventoryBuildResult:
    resolved_raw_folder = resolve_raw_folder(raw_folder)
    week_ending = parse_week_ending(resolved_raw_folder.name)
    inventory_source = find_inventory_source_file(resolved_raw_folder)
    output_path = (
        Path(output_file)
        if output_file
        else resolved_raw_folder / format_inventory_output_filename(week_ending)
    )
    removed_isbns_path = (
        resolved_raw_folder
        / format_legacy_bn_removed_isbns_filename(INVENTORY_PREFIX, week_ending)
    )

    if not output_file:
        previous_output = (
            resolved_raw_folder / format_previous_inventory_output_filename(week_ending)
        )
        if previous_output != output_path and previous_output.exists():
            previous_output.unlink()

    workbook = load_workbook(inventory_source)
    worksheet = workbook.active

    normalize_existing_headers(worksheet)
    remove_footer_rows(worksheet)
    last_data_row = get_last_data_row(worksheet)
    existing_index = build_existing_inventory_index(worksheet, last_data_row)
    pos_df = load_pos_dataframe(resolved_raw_folder)
    allowed_isbns = load_bn_upload_isbns()

    matched_updates = 0
    missing_rows = []
    for _, row in pos_df.iterrows():
        isbn = row["ISBN"]
        if isbn in existing_index:
            row_idx = existing_index[isbn]
            worksheet.cell(row=row_idx, column=7).value = row["OH_DC"]
            worksheet.cell(row=row_idx, column=9).value = row["OH_Stores"]
            worksheet.cell(row=row_idx, column=11).value = row["OO_DC"]
            worksheet.cell(row=row_idx, column=13).value = row["OO_Stores"]
            matched_updates += 1
        else:
            missing_rows.append(row)

    appended_rows = 0
    if missing_rows:
        missing_df = pd.DataFrame(missing_rows)
        style_row = (
            last_data_row
            if last_data_row >= INVENTORY_DATA_START_ROW
            else INVENTORY_DATA_START_ROW
        )
        appended_rows = append_missing_rows(
            worksheet,
            last_data_row + 1,
            missing_df,
            style_row,
        )
        last_data_row += appended_rows

    removed_rows = remove_disallowed_rows(worksheet, last_data_row, allowed_isbns)
    excluded_rows = len(removed_rows)
    last_data_row = get_last_data_row(worksheet)
    refresh_grand_total_row(worksheet, last_data_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    save_removed_isbns_report(removed_rows, removed_isbns_path)

    final_shape = (max(last_data_row - INVENTORY_DATA_START_ROW + 1, 0), 33)
    return InventoryBuildResult(
        raw_folder=resolved_raw_folder,
        source_file=inventory_source,
        output_file=output_path,
        removed_isbns_file=removed_isbns_path,
        matched_updates=matched_updates,
        appended_rows=appended_rows,
        excluded_rows=excluded_rows,
        final_shape=final_shape,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build the Barnes & Noble working Inventory file from Inventory*.xlsx and pos_combined."
    )
    parser.add_argument(
        "--raw-folder",
        help="Full path to a yyyy_mm_dd_raw_files folder. Defaults to the latest matching folder.",
    )
    parser.add_argument(
        "--output-file",
        help="Optional full output path for the working Inventory Excel file.",
    )
    return parser


def print_result_summary(result: InventoryBuildResult) -> None:
    print()
    print(f"Raw folder: {result.raw_folder}")
    print(f"Inventory source file: {result.source_file.name}")
    print(f"Matched ISBN overrides: {result.matched_updates:,}")
    print(f"Appended new ISBN rows: {result.appended_rows:,}")
    print(f"Rows removed by ISBN whitelist: {result.excluded_rows:,}")
    print(f"Final data shape: {result.final_shape}")
    print(f"Saved file: {result.output_file}")
    print(f"Removed ISBNs file: {result.removed_isbns_file}")


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    result = build_inventory_working_file(
        raw_folder=args.raw_folder,
        output_file=args.output_file,
    )
    print_result_summary(result)


if __name__ == "__main__":
    main()
