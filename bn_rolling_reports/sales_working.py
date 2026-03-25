from __future__ import annotations

import argparse
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from collections import OrderedDict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import (
    SALES_DATA_START_ROW,
    SALES_HEADER_ROW,
    SALES_PREFIX,
    SALES_REQUIRED_HEADERS,
    SALES_TOTAL_ROW,
    format_legacy_bn_removed_isbns_filename,
    format_legacy_bn_output_filename,
)
from isbn_utils import load_bn_upload_isbns, normalize_isbn, normalize_isbn_series
from pos_combiner import format_output_filename, parse_week_ending, resolve_raw_folder


FOOTER_MARKERS = (
    "Data available from",
    "Copyright",
)


@dataclass
class SalesBuildResult:
    raw_folder: Path
    source_file: Path
    output_file: Path
    removed_isbns_file: Path
    matched_updates: int
    appended_rows: int
    excluded_rows: int
    final_shape: tuple[int, int]


def format_sales_output_filename(week_ending: datetime) -> str:
    return format_legacy_bn_output_filename(SALES_PREFIX, week_ending)


def format_previous_sales_output_filename(week_ending: datetime) -> str:
    return f"{SALES_PREFIX}_{week_ending:%Y_%m_%d}.xlsx"


def find_sales_source_file(raw_folder: Path) -> Path:
    matches = sorted(
        [
            path
            for path in raw_folder.iterdir()
            if path.is_file()
            and path.suffix.lower() == ".xlsx"
            and path.name.lower().startswith("sales")
            and not path.name.startswith("~$")
        ],
        key=lambda path: path.name.lower(),
    )
    if not matches:
        raise FileNotFoundError(f"Could not find a Sales*.xlsx file in {raw_folder}")
    return matches[0]


def get_pos_output_file(raw_folder: Path) -> Path:
    week_ending = parse_week_ending(raw_folder.name)
    return raw_folder / format_output_filename(week_ending)


def load_pos_dataframe(raw_folder: Path) -> pd.DataFrame:
    pos_output = get_pos_output_file(raw_folder)
    if not pos_output.exists():
        raise FileNotFoundError(
            f"Combined POS file not found: {pos_output}. Run pos_combiner.py first."
        )

    df = pd.read_excel(pos_output, dtype={"ISBN": "string"})
    required = {"ISBN", "LW"}
    missing = required.difference(df.columns)
    if missing:
        raise ValueError(f"Combined POS file is missing required columns: {sorted(missing)}")

    df = df.loc[:, ["ISBN", "LW"]].copy()
    df["ISBN"] = normalize_isbn_series(df["ISBN"])
    df = df.loc[df["ISBN"].notna()].drop_duplicates(subset=["ISBN"], keep="first")
    return df


def normalize_existing_headers(ws) -> None:
    current_headers = [ws.cell(row=SALES_HEADER_ROW, column=idx).value for idx in range(1, 24)]
    if current_headers == SALES_REQUIRED_HEADERS:
        return

    for insert_at in (9, 13, 17, 21):
        ws.insert_cols(insert_at, amount=1)

    for col_idx, header in enumerate(SALES_REQUIRED_HEADERS, start=1):
        ws.cell(row=SALES_HEADER_ROW, column=col_idx).value = header


def align_sales_header_rows(ws) -> None:
    a5 = ws.cell(row=5, column=1).value
    a6 = ws.cell(row=6, column=1).value

    if a6 == "ISBN":
        return

    if a5 == "ISBN":
        ws.insert_rows(2, 1)
        ws.insert_rows(6, 1)
        ws.delete_rows(5, 1)
        return

    raise ValueError(
        "Unexpected Sales header layout. Expected 'ISBN' in A5 or A6 before processing."
    )


def remove_footer_rows(ws) -> None:
    for row_idx in range(SALES_DATA_START_ROW, ws.max_row + 1):
        first_value = ws.cell(row=row_idx, column=1).value
        if isinstance(first_value, str) and any(
            first_value.strip().startswith(marker) for marker in FOOTER_MARKERS
        ):
            ws.delete_rows(row_idx, ws.max_row - row_idx + 1)
            return


def get_last_data_row(ws) -> int:
    for row_idx in range(ws.max_row, SALES_DATA_START_ROW - 1, -1):
        row_has_data = any(
            ws.cell(row=row_idx, column=col_idx).value not in (None, "")
            for col_idx in range(1, 24)
        )
        if row_has_data:
            return row_idx
    return SALES_DATA_START_ROW - 1


def coerce_numeric_cell(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        return value


def coerce_fg_columns(ws, last_data_row: int) -> None:
    for row_idx in range(SALES_DATA_START_ROW, last_data_row + 1):
        for col_idx in (8, 10, 11):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = coerce_numeric_cell(cell.value)


def clone_row_styles(ws, source_row: int, target_row: int, max_col: int = 23) -> None:
    for col_idx in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col_idx)
        target_cell = ws.cell(row=target_row, column=col_idx)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def build_existing_sales_index(ws, last_data_row: int) -> dict[str, int]:
    index: dict[str, int] = {}
    for row_idx in range(SALES_DATA_START_ROW, last_data_row + 1):
        raw_isbn = ws.cell(row=row_idx, column=1).value
        isbn = normalize_isbn(raw_isbn)
        if not isbn or isbn in index:
            continue
        index[isbn] = row_idx
        ws.cell(row=row_idx, column=1).value = isbn
        ws.cell(row=row_idx, column=1).number_format = "@"
    return index


def consolidate_sales_rows(ws, last_data_row: int) -> int:
    aggregated: OrderedDict[str, dict[str, object]] = OrderedDict()

    for row_idx in range(SALES_DATA_START_ROW, last_data_row + 1):
        isbn = normalize_isbn(ws.cell(row=row_idx, column=1).value)
        if not isbn:
            continue

        subject_code = ws.cell(row=row_idx, column=6).value
        dept_code = ws.cell(row=row_idx, column=7).value
        total_value = coerce_numeric_cell(ws.cell(row=row_idx, column=8).value)
        bn_com_value = coerce_numeric_cell(ws.cell(row=row_idx, column=10).value)
        bn_value = coerce_numeric_cell(ws.cell(row=row_idx, column=11).value)

        if isbn not in aggregated:
            aggregated[isbn] = {
                "subject_code": subject_code,
                "dept_code": dept_code,
                "total_lw": 0 if total_value in (None, "") else total_value,
                "bn_com_lw": 0 if bn_com_value in (None, "") else bn_com_value,
                "bn_lw": 0 if bn_value in (None, "") else bn_value,
            }
            continue

        if not aggregated[isbn]["subject_code"] and subject_code not in (None, ""):
            aggregated[isbn]["subject_code"] = subject_code
        if not aggregated[isbn]["dept_code"] and dept_code not in (None, ""):
            aggregated[isbn]["dept_code"] = dept_code
        if total_value not in (None, ""):
            aggregated[isbn]["total_lw"] += total_value
        if bn_com_value not in (None, ""):
            aggregated[isbn]["bn_com_lw"] += bn_com_value
        if bn_value not in (None, ""):
            aggregated[isbn]["bn_lw"] += bn_value

    row_values = list(aggregated.items())
    target_last_row = SALES_DATA_START_ROW + len(row_values) - 1

    for offset, (isbn, values) in enumerate(row_values):
        row_idx = SALES_DATA_START_ROW + offset
        for col_idx in range(1, 24):
            ws.cell(row=row_idx, column=col_idx).value = None
        ws.cell(row=row_idx, column=1).value = isbn
        ws.cell(row=row_idx, column=1).number_format = "@"
        ws.cell(row=row_idx, column=6).value = values["subject_code"]
        ws.cell(row=row_idx, column=7).value = values["dept_code"]
        ws.cell(row=row_idx, column=8).value = values["total_lw"]
        ws.cell(row=row_idx, column=9).value = None
        ws.cell(row=row_idx, column=10).value = values["bn_com_lw"]
        ws.cell(row=row_idx, column=11).value = values["bn_lw"]

    if target_last_row < last_data_row:
        ws.delete_rows(target_last_row + 1, last_data_row - target_last_row)

    return max(target_last_row, SALES_DATA_START_ROW - 1)


def excel_safe_value(value):
    return None if pd.isna(value) else value


def append_missing_rows(ws, start_row: int, pos_missing: pd.DataFrame, style_row: int) -> int:
    appended = 0
    for _, row in pos_missing.iterrows():
        target_row = start_row + appended
        clone_row_styles(ws, style_row, target_row)
        for col_idx in range(1, 24):
            ws.cell(row=target_row, column=col_idx).value = None
        ws.cell(row=target_row, column=1).value = excel_safe_value(row["ISBN"])
        ws.cell(row=target_row, column=1).number_format = "@"
        ws.cell(row=target_row, column=8).value = excel_safe_value(row["LW"])
        appended += 1
    return appended


def remove_disallowed_rows(ws, last_data_row: int, allowed_isbns: set[str]) -> int:
    removed_rows: list[dict[str, object]] = []
    for row_idx in range(last_data_row, SALES_DATA_START_ROW - 1, -1):
        isbn = normalize_isbn(ws.cell(row=row_idx, column=1).value)
        if not isbn or isbn not in allowed_isbns:
            removed_rows.append(
                {
                    "ISBN": isbn or "",
                    "Title": ws.cell(row=row_idx, column=2).value,
                    "Author": ws.cell(row=row_idx, column=4).value,
                    "Imprint": ws.cell(row=row_idx, column=3).value,
                    "Reason": "Missing/invalid ISBN" if not isbn else "Not in BN upload whitelist",
                }
            )
            ws.delete_rows(row_idx, 1)
    removed_rows.reverse()
    return removed_rows


def save_removed_isbns_report(removed_rows: list[dict[str, object]], output_file: Path) -> None:
    df = pd.DataFrame(
        removed_rows,
        columns=["ISBN", "Title", "Author", "Imprint", "Reason"],
    )
    output_file.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_file, index=False)


def refresh_grand_total_row(ws, last_data_row: int) -> None:
    if last_data_row < SALES_DATA_START_ROW:
        ws.cell(row=SALES_TOTAL_ROW, column=8).value = 0
        ws.cell(row=SALES_TOTAL_ROW, column=9).value = 0
        ws.cell(row=SALES_TOTAL_ROW, column=10).value = 0
        ws.cell(row=SALES_TOTAL_ROW, column=11).value = 0
        ws.cell(row=SALES_TOTAL_ROW, column=16).value = None
        return

    ws.cell(row=SALES_TOTAL_ROW, column=9).value = 0
    for col_idx in (8, 10, 11):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=SALES_TOTAL_ROW, column=col_idx).value = (
            f"=SUM({col_letter}{SALES_DATA_START_ROW}:{col_letter}{last_data_row})"
        )
    ws.cell(row=SALES_TOTAL_ROW, column=16).value = None


def build_sales_working_file(
    raw_folder: str | Path | None = None,
    output_file: str | Path | None = None,
) -> SalesBuildResult:
    resolved_raw_folder = resolve_raw_folder(raw_folder)
    week_ending = parse_week_ending(resolved_raw_folder.name)
    sales_source = find_sales_source_file(resolved_raw_folder)
    output_path = (
        Path(output_file)
        if output_file
        else resolved_raw_folder / format_sales_output_filename(week_ending)
    )
    removed_isbns_path = (
        resolved_raw_folder / format_legacy_bn_removed_isbns_filename(SALES_PREFIX, week_ending)
    )

    if not output_file:
        previous_output = resolved_raw_folder / format_previous_sales_output_filename(week_ending)
        if previous_output != output_path and previous_output.exists():
            previous_output.unlink()

    workbook = load_workbook(sales_source)
    worksheet = workbook.active

    align_sales_header_rows(worksheet)
    normalize_existing_headers(worksheet)
    remove_footer_rows(worksheet)
    last_data_row = get_last_data_row(worksheet)
    coerce_fg_columns(worksheet, last_data_row)
    last_data_row = consolidate_sales_rows(worksheet, last_data_row)

    existing_index = build_existing_sales_index(worksheet, last_data_row)
    pos_df = load_pos_dataframe(resolved_raw_folder)
    allowed_isbns = load_bn_upload_isbns()

    matched_updates = 0
    missing_rows = []
    for _, row in pos_df.iterrows():
        isbn = row["ISBN"]
        if isbn in existing_index:
            row_idx = existing_index[isbn]
            worksheet.cell(row=row_idx, column=8).value = row["LW"]
            matched_updates += 1
        else:
            missing_rows.append(row)

    appended_rows = 0
    if missing_rows:
        missing_df = pd.DataFrame(missing_rows)
        style_row = last_data_row if last_data_row >= SALES_DATA_START_ROW else SALES_DATA_START_ROW
        appended_rows = append_missing_rows(
            worksheet,
            last_data_row + 1,
            missing_df,
            style_row,
        )
        last_data_row += appended_rows

    removed_rows = remove_disallowed_rows(worksheet, last_data_row, allowed_isbns)
    excluded_rows = len(removed_rows)
    last_data_row = get_last_data_row(worksheet)
    refresh_grand_total_row(worksheet, last_data_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    save_removed_isbns_report(removed_rows, removed_isbns_path)

    final_shape = (max(last_data_row - SALES_DATA_START_ROW + 1, 0), 23)
    return SalesBuildResult(
        raw_folder=resolved_raw_folder,
        source_file=sales_source,
        output_file=output_path,
        removed_isbns_file=removed_isbns_path,
        matched_updates=matched_updates,
        appended_rows=appended_rows,
        excluded_rows=excluded_rows,
        final_shape=final_shape,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build the Barnes & Noble working Sales file from Sales*.xlsx and pos_combined."
    )
    parser.add_argument(
        "--raw-folder",
        help="Full path to a yyyy_mm_dd_raw_files folder. Defaults to the latest matching folder.",
    )
    parser.add_argument(
        "--output-file",
        help="Optional full output path for the working Sales Excel file.",
    )
    return parser


def print_result_summary(result: SalesBuildResult) -> None:
    print()
    print(f"Raw folder: {result.raw_folder}")
    print(f"Sales source file: {result.source_file.name}")
    print(f"Matched ISBN overrides: {result.matched_updates:,}")
    print(f"Appended new ISBN rows: {result.appended_rows:,}")
    print(f"Rows removed by ISBN whitelist: {result.excluded_rows:,}")
    print(f"Final data shape: {result.final_shape}")
    print(f"Saved file: {result.output_file}")
    print(f"Removed ISBNs file: {result.removed_isbns_file}")


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    result = build_sales_working_file(raw_folder=args.raw_folder, output_file=args.output_file)
    print_result_summary(result)


if __name__ == "__main__":
    main()
