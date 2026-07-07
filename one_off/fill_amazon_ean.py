from __future__ import annotations

import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
AMAZON_SQL_UPLOAD_DIR = REPO_ROOT / "amazon_sql_upload"

for path in [REPO_ROOT, AMAZON_SQL_UPLOAD_DIR]:
    path_text = str(path)
    if path_text not in sys.path:
        sys.path.insert(0, path_text)

from asin_manual_key import asin_isbn_manual_key  # noqa: E402
from load_catalog import df_catalog  # noqa: E402
from load_ebs_isbn_key import isbn_key  # noqa: E402
from load_ypticod import load_ypticod  # noqa: E402


SOURCE_DIR = REPO_ROOT / "one_off"
OUTPUT_FILE = SOURCE_DIR / "Amazon_consolidated_with_EAN.xlsx"
MISSING_FILE = SOURCE_DIR / "Amazon_missing_asins.xlsx"
AMOUNT_SUFFIX_RE = re.compile(r"\s+\$?\d[\d,]*\.\d{2}$")


def normalize_asin(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip().upper().zfill(10)


def normalize_isbn(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    digits = "".join(char for char in str(value).strip() if char.isdigit())
    if not digits:
        return ""
    if len(digits) < 13:
        return digits.zfill(13)
    if len(digits) > 13 and digits.startswith("0"):
        return digits[-13:]
    return digits[:13]


def source_label(path: Path) -> str:
    return AMOUNT_SUFFIX_RE.sub("", path.stem).strip()


def source_files() -> list[Path]:
    return sorted(
        path
        for path in SOURCE_DIR.glob("*.xlsx")
        if not path.name.startswith("~$")
        and path.name not in {OUTPUT_FILE.name, MISSING_FILE.name}
    )


def load_asins(files: list[Path]) -> set[str]:
    asins: set[str] = set()
    for path in files:
        df = pd.read_excel(path, dtype=str, usecols=["Asin"])
        asins.update(df["Asin"].map(normalize_asin).dropna())
    asins.discard("")
    return asins


def load_mapping() -> pd.DataFrame:
    df_ypticod = load_ypticod()
    df_catalog = df_catalog_safe()
    df_isbn = isbn_key()

    for frame in [df_ypticod, df_catalog]:
        frame["ASIN"] = frame["ASIN"].map(normalize_asin)

    df_isbn = df_isbn.copy()
    df_isbn["ISBN"] = df_isbn["ISBN"].map(normalize_isbn)
    isbn_set = set(df_isbn["ISBN"].dropna().unique())
    isbn_set.discard("")

    mapping = df_ypticod.copy()
    mapping["ISBN"] = mapping["ISBN"].map(normalize_isbn)
    mapping = mapping.drop_duplicates(subset=["ASIN"], keep="first")

    catalog = df_catalog.drop_duplicates(subset=["ASIN"], keep="first")
    combined = pd.DataFrame({"ASIN": sorted(set(mapping["ASIN"]) | set(catalog["ASIN"]))})
    combined = combined.merge(mapping.rename(columns={"ISBN": "ISBN_ypticod"}), on="ASIN", how="left")
    combined = combined.merge(catalog, on="ASIN", how="left")

    isbn_col = combined["ISBN_ypticod"].fillna("").map(normalize_isbn)
    for column in ["EAN", "ISBN_Amz", "Model Number"]:
        values = combined[column].fillna("").map(normalize_isbn)
        mask = (isbn_col == "") & values.isin(isbn_set)
        isbn_col = pd.Series(np.where(mask, values, isbn_col), index=combined.index)

    combined["ISBN"] = isbn_col
    combined["ISBN"] = combined.apply(
        lambda row: asin_isbn_manual_key.get(row["ASIN"], row["ISBN"]),
        axis=1,
    )
    combined = combined[combined["ISBN"] != ""]
    return combined[["ASIN", "ISBN"]].drop_duplicates(subset=["ASIN"], keep="first")


def df_catalog_safe() -> pd.DataFrame:
    catalog = df_catalog()
    for column in ["ASIN", "EAN", "ISBN_Amz", "Model Number"]:
        if column not in catalog.columns:
            catalog[column] = ""
    return catalog[["ASIN", "EAN", "ISBN_Amz", "Model Number"]].copy()


def update_workbook_ean(path: Path, asin_to_isbn: dict[str, str]) -> int:
    workbook = load_workbook(path)
    sheet = workbook.active
    headers = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value is not None}
    if "Asin" not in headers or "EAN" not in headers:
        raise ValueError(f"{path.name} is missing Asin or EAN header.")

    asin_col = headers["Asin"]
    ean_col = headers["EAN"]
    filled = 0
    for row in range(2, sheet.max_row + 1):
        asin = normalize_asin(sheet.cell(row=row, column=asin_col).value)
        isbn = asin_to_isbn.get(asin, "")
        if isbn:
            sheet.cell(row=row, column=ean_col).value = isbn
            filled += 1
    workbook.save(path)
    return filled


def read_completed_file(path: Path, asin_to_isbn: dict[str, str]) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=str)
    df.insert(0, "Source File", source_label(path))
    df["Asin"] = df["Asin"].map(normalize_asin)
    df["EAN"] = df["Asin"].map(asin_to_isbn).fillna(df["EAN"].fillna(""))
    return df


def main() -> None:
    files = source_files()
    if not files:
        raise FileNotFoundError(f"No source xlsx files found in {SOURCE_DIR}")

    source_asins = load_asins(files)
    mapping = load_mapping()
    mapping = mapping[mapping["ASIN"].isin(source_asins)].copy()
    asin_to_isbn = dict(zip(mapping["ASIN"], mapping["ISBN"]))

    combined_frames = []
    file_summary = []
    for path in files:
        filled = update_workbook_ean(path, asin_to_isbn)
        completed = read_completed_file(path, asin_to_isbn)
        combined_frames.append(completed)
        file_summary.append(
            {
                "Source File": source_label(path),
                "Rows": len(completed),
                "EAN Filled Rows": filled,
                "Unique ASINs": completed["Asin"].nunique(dropna=True),
            }
        )

    combined = pd.concat(combined_frames, ignore_index=True)
    missing_asins = sorted(set(combined.loc[combined["EAN"].fillna("") == "", "Asin"]) - {""})
    missing = pd.DataFrame({"ASIN": missing_asins})
    summary = pd.DataFrame(file_summary)

    with pd.ExcelWriter(OUTPUT_FILE, engine="xlsxwriter") as writer:
        combined.to_excel(writer, sheet_name="Consolidated", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        missing.to_excel(writer, sheet_name="Missing ASINs", index=False)

    missing.to_excel(MISSING_FILE, index=False)

    print(f"Source files processed: {len(files):,}")
    print(f"Combined rows: {len(combined):,}")
    print(f"Unique source ASINs: {len(source_asins):,}")
    print(f"Mapped ASINs: {len(mapping):,}")
    print(f"Missing ASINs: {len(missing):,}")
    print(f"Wrote: {OUTPUT_FILE}")
    print(f"Wrote: {MISSING_FILE}")


if __name__ == "__main__":
    main()
