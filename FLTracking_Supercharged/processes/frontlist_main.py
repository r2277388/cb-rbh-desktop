import argparse
from dataclasses import dataclass
from datetime import datetime
from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path
from shutil import copy2
import sys
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string


sys.path.append(str(Path(__file__).resolve().parents[1]))

from config import FRONTLIST_MAIN_OUTPUT_DIR
from isbn_utils import normalize_isbn, normalize_isbn_series
from processes.amazon_preorders import (
    AMAZON_PREORDERS_PATH,
    load_amazon_preorders,
    load_amazon_preorders_cached,
)
from processes.amazon_sellthrough import get_amazon_sellthrough_source_metadata
from processes.amazon_sellthrough import load_amazon_sellthrough
from processes.barnes_noble_weekly import (
    get_barnes_noble_source_metadata,
    load_barnes_noble_weekly_cached,
)
from processes.faire_orders import SQL_FILE as FAIRE_ORDERS_SQL_FILE
from processes.faire_orders import load_faire_orders
from processes.faire_qty import SQL_FILE as FAIRE_QTY_SQL_FILE
from processes.faire_qty import load_faire_qty
from processes.bookshop_preorders import (
    load_bookshop_preorders_cached,
    parse_bookshop_report_date,
    resolve_bookshop_preorders_path,
)
from processes.ingram_daily_report import (
    build_modified_date_header,
    load_ingram_daily_report,
    load_ingram_daily_report_cached,
    resolve_ingram_daily_report_path,
)
from processes.inventory_detail import (
    load_inventory_detail,
    load_inventory_detail_cached,
    resolve_inventory_detail_path,
)


def _load_shared_paths():
    shared_path = Path(__file__).resolve().parents[2] / "paths" / "process_paths.py"
    spec = spec_from_file_location("_shared_process_paths", shared_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Unable to load shared process paths from {shared_path}")
    module = module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


FRONTLIST_DIR = _load_shared_paths().FRONTLIST_TRACKING_FOLDER
FRONTLIST_MAIN_ROW1_LABELS = [
    "Soonest Reprint Date",
    "Soonest Reprint Qty",
    "Avail to Sell",
    "Current Stock Freeze Total",
    "Future Reserve Qty Total",
    "Amazon Preorders",
    "Bookshop Preorders",
    "Sellthru Amazon (LTD)",
    "OH Amazon",
    "Faire LTD Sales",
    "Faire OO",
    "Last 4 Weeks Ingram",
    "OH Ingram",
    "Ingram Preorders",
    "OH BN Superstores",
    "OH BN DC",
    "Sellthru B&N (LTD)",
]

SUMMARY_INSERT_AT_COLUMN = "N"
SUMMARY_HEADER_ROW = 6
SUMMARY_FIRST_DATA_ROW = 7
SUMMARY_HEADER_TOP_ROW = 2
SUPERCHARGED_SUFFIX = "_SuperCharged"
SUPERCHARGED_TOP_BAND_FILL = "222B35"
SUPERCHARGED_UPPER_PANEL_FILL = "9BC2E6"


@dataclass(frozen=True)
class SuperchargedColumnSpec:
    excel_column: str
    header_label: str
    source_column: str | None
    value_kind: str
    width: float
    row5_date_rule: str | None = None


@dataclass(frozen=True)
class SuperchargedSectionSpec:
    start_column: str
    title: str
    fill_tint: float


SUPERCHARGED_SECTIONS = [
    SuperchargedSectionSpec("N", "Inventory", 0.5999938962981048),
    SuperchargedSectionSpec("S", "Pre-orders by Account as of", 0.3999755851924192),
    SuperchargedSectionSpec("X", "Sellthrough by Account as of", 0.7999816888943144),
    SuperchargedSectionSpec("AB", "Account Inventory as of", 0.3999755851924192),
]

SUPERCHARGED_COLUMNS = [
    SuperchargedColumnSpec("N", "Avail to Sell", "Available To Sell", "number", 13.0),
    SuperchargedColumnSpec("O", "Current Stock Freeze Total", "Frozen", "number", 13.0),
    SuperchargedColumnSpec("P", "Soonest Reprint Date", "Reprint Due Date", "date", 13.85546875),
    SuperchargedColumnSpec("Q", "Soonest Reprint Qty", "Reprint Quantity", "number", 8.85546875),
    SuperchargedColumnSpec("R", "Future Reserve Qty Total", "Reprint Freeze", "number", 13.0),
    SuperchargedColumnSpec("S", "Amazon Preorders", "AmzPreOrders", "number", 9.5703125, "amazon_preorders_modified"),
    SuperchargedColumnSpec("T", "BN Preorders", None, "number", 10.7109375),
    SuperchargedColumnSpec("U", "Bookshop Preorders", "BookshopPreOrders", "number", 13.0, "bookshop_report"),
    SuperchargedColumnSpec("V", "Ingram Preorders", "IngramPreOrders", "number", 13.0, "ingram_modified"),
    SuperchargedColumnSpec("W", "Faire OO", "Faire_OO_qty", "number", 8.85546875, "run_date"),
    SuperchargedColumnSpec("X", "Sellthru Amazon (LTD)", "AmzUnitShipped_LTD", "number", 13.0, "amazon_sellthrough_report"),
    SuperchargedColumnSpec("Y", "Sellthru B&N (LTD)", "B&N_LTD", "number", 13.0, "bn_report"),
    SuperchargedColumnSpec("Z", "Last 4 Weeks Ingram", "Ingram4WkSales", "number", 13.0, "ingram_modified"),
    SuperchargedColumnSpec("AA", "Faire LTD Sales", "FaireQty", "number", 13.0, "run_date"),
    SuperchargedColumnSpec("AB", "OH Amazon", "AmzOnHand", "number", 13.0, "amazon_sellthrough_report"),
    SuperchargedColumnSpec("AC", "OH BN Superstores", "B&N_OH_Store", "number", 13.0, "bn_report"),
    SuperchargedColumnSpec("AD", "OH BN DC", "B&N_OH_DC", "number", 13.0, "bn_report"),
    SuperchargedColumnSpec("AE", "OH Ingram", "IngramOH", "number", 13.0, "ingram_modified"),
]


def resolve_frontlist_tracking_path(source_dir: Path = FRONTLIST_DIR) -> Path:
    files = [
        path
        for path in source_dir.glob("*.xlsx")
        if not path.name.startswith("~$")
        and SUPERCHARGED_SUFFIX.lower() not in path.stem.lower()
    ]
    if not files:
        raise FileNotFoundError(f"No Frontlist Tracking workbook found in: {source_dir}")
    return max(files, key=lambda path: path.stat().st_mtime)


def _validate_frontlist_source_path(frontlist_path: Path) -> Path:
    if SUPERCHARGED_SUFFIX.lower() in frontlist_path.stem.lower():
        raise ValueError(
            "Frontlist source workbook cannot be a _SuperCharged file. "
            f"Choose the original Frontlist Tracking workbook instead: {frontlist_path}"
        )
    return frontlist_path


def load_frontlist_isbns(source_path: Path | None = None) -> pd.DataFrame:
    resolved_path = _validate_frontlist_source_path(source_path) if source_path else resolve_frontlist_tracking_path()
    df = pd.read_excel(
        resolved_path,
        header=5,
        usecols=["ISBN-13"],
        dtype={"ISBN-13": "object"},
        engine="openpyxl",
    )

    result = df.rename(columns={"ISBN-13": "ISBN"}).copy()
    result["ISBN"] = normalize_isbn_series(result["ISBN"])
    result = result.dropna(subset=["ISBN"]).drop_duplicates(subset=["ISBN"]).reset_index(drop=True)
    return result


def dedupe_source_on_isbn(df: pd.DataFrame) -> pd.DataFrame:
    if df["ISBN"].is_unique:
        return df

    aggregations: dict[str, str] = {}
    for col in df.columns:
        if col == "ISBN":
            continue
        if pd.api.types.is_numeric_dtype(df[col]):
            aggregations[col] = "sum"
        else:
            aggregations[col] = "first"

    return df.groupby("ISBN", as_index=False).agg(aggregations)


def fill_missing_metric_values(df: pd.DataFrame) -> pd.DataFrame:
    protected_cols = {"ISBN", "Reprint Due Date"}
    protected_prefixes = ("IngramUpdated_", "BNUpdated_")

    for col in df.columns:
        if col in protected_cols or col.startswith(protected_prefixes):
            continue

        converted = pd.to_numeric(df[col], errors="coerce")
        original_non_null = df[col].notna().sum()
        converted_non_null = converted.notna().sum()

        if original_non_null == converted_non_null:
            filled = converted.fillna(0)
            if (filled % 1 == 0).all():
                df[col] = filled.astype("Int64")
            else:
                df[col] = filled

    return df


def build_metadata_sheet(frontlist_path: Path, combined: pd.DataFrame) -> pd.DataFrame:
    ingram_path = resolve_ingram_daily_report_path()
    barnes_noble_metadata = get_barnes_noble_source_metadata()
    amazon_sellthrough_metadata = get_amazon_sellthrough_source_metadata()
    inventory_detail_path = resolve_inventory_detail_path()
    amazon_preorders_path = AMAZON_PREORDERS_PATH
    bookshop_preorders_path = resolve_bookshop_preorders_path()
    bookshop_report_date = parse_bookshop_report_date(bookshop_preorders_path)
    run_date = datetime.now().strftime("%m/%d/%Y")

    _, ingram_report_date = build_modified_date_header(ingram_path)

    amz_last_week_value = ""
    if "AmzLastWeek" in combined.columns:
        non_null = combined["AmzLastWeek"].dropna()
        if not non_null.empty:
            amz_last_week_value = pd.to_datetime(non_null.iloc[0]).strftime("%m/%d/%Y")

    rows = [
        {
            "Source": "Frontlist Tracking",
            "FileName": frontlist_path.name,
            "ReportDate": "",
            "ModifiedDate": datetime.fromtimestamp(frontlist_path.stat().st_mtime).strftime("%m/%d/%Y"),
        },
        {
            "Source": "Ingram",
            "FileName": ingram_path.name,
            "ReportDate": ingram_report_date,
            "ModifiedDate": datetime.fromtimestamp(ingram_path.stat().st_mtime).strftime("%m/%d/%Y"),
        },
        {
            "Source": "Barnes & Noble",
            "FileName": _metadata_filename(barnes_noble_metadata),
            "ReportDate": barnes_noble_metadata["report_date"],
            "ModifiedDate": barnes_noble_metadata["modified_date"],
        },
        {
            "Source": "Inventory Detail",
            "FileName": inventory_detail_path.name,
            "ReportDate": "",
            "ModifiedDate": datetime.fromtimestamp(inventory_detail_path.stat().st_mtime).strftime("%m/%d/%Y"),
        },
        {
            "Source": "Amazon Preorders",
            "FileName": amazon_preorders_path.name,
            "ReportDate": "",
            "ModifiedDate": datetime.fromtimestamp(amazon_preorders_path.stat().st_mtime).strftime("%m/%d/%Y"),
        },
        {
            "Source": "Bookshop Preorders",
            "FileName": bookshop_preorders_path.name,
            "ReportDate": bookshop_report_date,
            "ModifiedDate": datetime.fromtimestamp(bookshop_preorders_path.stat().st_mtime).strftime("%m/%d/%Y"),
        },
        {
            "Source": "Amazon Sellthrough",
            "FileName": _metadata_filename(amazon_sellthrough_metadata),
            "ReportDate": amz_last_week_value or amazon_sellthrough_metadata["report_date"],
            "ModifiedDate": amazon_sellthrough_metadata["modified_date"],
        },
        {
            "Source": "Faire Qty SQL",
            "FileName": FAIRE_QTY_SQL_FILE.name,
            "ReportDate": "",
            "ModifiedDate": run_date,
        },
        {
            "Source": "Faire Orders SQL",
            "FileName": FAIRE_ORDERS_SQL_FILE.name,
            "ReportDate": "",
            "ModifiedDate": run_date,
        },
    ]

    return pd.DataFrame(rows)


def _metadata_filename(metadata: dict[str, object]) -> str:
    if "cache_path" in metadata:
        return Path(metadata["cache_path"]).name
    if "sql_path" in metadata:
        return Path(metadata["sql_path"]).name
    if "source_path" in metadata:
        return Path(metadata["source_path"]).name
    if "sales_path" in metadata and "inventory_path" in metadata:
        return f"{Path(metadata['sales_path']).name}; {Path(metadata['inventory_path']).name}"
    return ""


def build_frontlist_main(frontlist_path: Path | None = None) -> tuple[pd.DataFrame, pd.DataFrame, Path]:
    resolved_frontlist_path = (
        _validate_frontlist_source_path(frontlist_path) if frontlist_path else resolve_frontlist_tracking_path()
    )
    combined = load_frontlist_isbns(resolved_frontlist_path)

    source_frames = [
        load_inventory_detail_cached()[0],
        load_amazon_preorders_cached()[0],
        load_bookshop_preorders_cached()[0],
        load_amazon_sellthrough(),
        load_faire_qty(),
        load_faire_orders(),
        load_ingram_daily_report_cached()[0],
        load_barnes_noble_weekly_cached()[0],
    ]

    for source_df in source_frames:
        combined = combined.merge(dedupe_source_on_isbn(source_df), on="ISBN", how="left")

    metadata_df = build_metadata_sheet(resolved_frontlist_path, combined)
    drop_cols = [
        col for col in combined.columns
        if col == "AmzLastWeek" or col.startswith("IngramUpdated_") or col.startswith("BNUpdated_")
    ]
    if drop_cols:
        combined = combined.drop(columns=drop_cols)

    combined = fill_missing_metric_values(combined)

    return combined, metadata_df, resolved_frontlist_path


def save_frontlist_main_output(
    df: pd.DataFrame,
    metadata_df: pd.DataFrame,
    as_of: datetime | None = None,
    output_dir: Path | None = None,
) -> Path:
    resolved_output_dir = output_dir or FRONTLIST_MAIN_OUTPUT_DIR
    resolved_output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = as_of or datetime.now()
    output_path = resolved_output_dir / f"frontlist_main_{timestamp.strftime('%Y_%m_%d')}.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="FrontlistMain", index=False, startrow=2)
        metadata_df.to_excel(writer, sheet_name="SourceDates", index=False)
        worksheet = writer.book["FrontlistMain"]
        for col_idx, label in enumerate(FRONTLIST_MAIN_ROW1_LABELS, start=2):
            worksheet.cell(row=1, column=col_idx).value = label
    return output_path


def _normalize_isbn_value(value: object) -> str | None:
    normalized = normalize_isbn(value, set())
    if normalized is None:
        return None
    return str(normalized)


def _build_supercharged_output_path(frontlist_path: Path) -> Path:
    return frontlist_path.with_name(f"{frontlist_path.stem}{SUPERCHARGED_SUFFIX}{frontlist_path.suffix}")


def _resolve_supercharged_output_path(
    frontlist_path: Path,
    output_dir: Path | None = None,
) -> Path:
    default_output_path = _build_supercharged_output_path(frontlist_path)
    if output_dir is None:
        return default_output_path
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / default_output_path.name


def _find_column_by_header(worksheet, header_name: str, row: int = SUMMARY_HEADER_ROW) -> int:
    for col_idx in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=row, column=col_idx).value == header_name:
            return col_idx
    raise ValueError(f"Could not find '{header_name}' on row {row} of the Summary sheet.")


def _copy_cell_style(target_cell, template_cell, *, alignment: Alignment | None = None) -> None:
    target_cell._style = copy(template_cell._style)
    target_cell.font = copy(template_cell.font)
    target_cell.fill = copy(template_cell.fill)
    target_cell.border = copy(template_cell.border)
    target_cell.number_format = template_cell.number_format
    target_cell.protection = copy(template_cell.protection)
    target_cell.alignment = copy(alignment if alignment is not None else template_cell.alignment)


def _find_supercharged_section(column_letter: str) -> SuperchargedSectionSpec:
    column_idx = column_index_from_string(column_letter)
    current_section = SUPERCHARGED_SECTIONS[0]
    for section in SUPERCHARGED_SECTIONS:
        if column_index_from_string(section.start_column) <= column_idx:
            current_section = section
        else:
            break
    return current_section


def _build_section_fill(section: SuperchargedSectionSpec) -> PatternFill:
    return PatternFill(
        fill_type="solid",
        fgColor=Color(theme=8, tint=section.fill_tint),
    )


def _apply_supercharged_upper_fill(worksheet) -> None:
    dark_fill = PatternFill(fill_type="solid", fgColor=SUPERCHARGED_TOP_BAND_FILL)
    light_fill = PatternFill(fill_type="solid", fgColor=SUPERCHARGED_UPPER_PANEL_FILL)

    for row_idx in range(1, 4):
        fill = dark_fill if row_idx == 1 else light_fill
        for spec in SUPERCHARGED_COLUMNS:
            worksheet[f"{spec.excel_column}{row_idx}"].fill = copy(fill)


def _build_supercharged_header_border(
    *,
    row_idx: int,
    is_block_start: bool,
) -> Border:
    medium_side = Side(style="medium")
    return Border(
        left=medium_side if is_block_start else Side(),
        top=medium_side if row_idx == 4 else Side(),
        bottom=medium_side if row_idx == 6 else Side(),
        right=Side(),
    )


def _apply_builtin_supercharged_header_style(cell, *, row_idx: int, section: SuperchargedSectionSpec) -> None:
    is_block_start = cell.column == column_index_from_string(section.start_column)
    cell.font = Font(name="Calibri", sz=11, bold=True, color="FF000000")
    cell.fill = _build_section_fill(section)
    cell.border = _build_supercharged_header_border(row_idx=row_idx, is_block_start=is_block_start)
    cell.protection = copy(cell.protection)

    if row_idx == 4:
        cell.alignment = Alignment(horizontal="centerContinuous")
        if section.start_column == "N":
            cell.number_format = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"_);_(@_)'
        else:
            cell.number_format = "# ?/?"
    elif row_idx == 5:
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = "mm/dd/yy;@" if cell.value else "General"
    else:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.number_format = "0" if cell.column == column_index_from_string("AE") else "General"


def _get_overlapping_merged_ranges(worksheet, min_col: int, max_col: int, min_row: int, max_row: int):
    ranges_to_unmerge = []
    for merged_range in worksheet.merged_cells.ranges:
        overlaps_columns = merged_range.min_col <= max_col and merged_range.max_col >= min_col
        overlaps_rows = merged_range.min_row <= max_row and merged_range.max_row >= min_row
        if overlaps_columns and overlaps_rows:
            ranges_to_unmerge.append(merged_range)
    return ranges_to_unmerge


def _clear_overlapping_merged_ranges(worksheet, min_col: int, max_col: int, min_row: int, max_row: int) -> None:
    ranges_to_unmerge = _get_overlapping_merged_ranges(worksheet, min_col, max_col, min_row, max_row)

    for merged_range in ranges_to_unmerge:
        worksheet.unmerge_cells(str(merged_range))


def _restore_shifted_merged_ranges(worksheet, merged_ranges, *, insert_col_idx: int, inserted_width: int) -> None:
    net_shift = inserted_width - 1
    for merged_range in merged_ranges:
        if merged_range.min_col < insert_col_idx:
            continue
        new_min_col = merged_range.min_col + net_shift
        new_max_col = merged_range.max_col + net_shift
        worksheet.merge_cells(
            start_row=merged_range.min_row,
            start_column=new_min_col,
            end_row=merged_range.max_row,
            end_column=new_max_col,
        )


def _format_short_date(value: object) -> str | None:
    if value in (None, "", pd.NaT):
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.strftime("%m/%d")


def _sort_summary_rows_by_column_desc(
    worksheet,
    *,
    sort_column_letter: str,
    first_data_row: int,
) -> None:
    sort_col_idx = column_index_from_string(sort_column_letter)
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    row_records = []
    for row_idx in range(first_data_row, max_row + 1):
        values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
        sort_value = worksheet.cell(row=row_idx, column=sort_col_idx).value
        numeric_sort_value = pd.to_numeric(sort_value, errors="coerce")
        row_records.append(
            (
                float("-inf") if pd.isna(numeric_sort_value) else float(numeric_sort_value),
                row_idx,
                values,
            )
        )

    row_records.sort(key=lambda record: (record[0], record[1]), reverse=True)

    for target_row_idx, (_, _, values) in enumerate(row_records, start=first_data_row):
        for col_idx, value in enumerate(values, start=1):
            worksheet.cell(row=target_row_idx, column=col_idx).value = value


def _build_summary_row5_dates(metadata_df: pd.DataFrame) -> dict[str, str | None]:
    metadata_by_source = metadata_df.set_index("Source").to_dict("index")
    run_date = datetime.now().strftime("%m/%d")

    def metadata_date(source: str, key: str) -> str | None:
        row = metadata_by_source.get(source)
        if not row:
            return None
        return _format_short_date(row.get(key))

    date_rule_values = {
        "amazon_preorders_modified": metadata_date("Amazon Preorders", "ModifiedDate"),
        "bookshop_report": metadata_date("Bookshop Preorders", "ReportDate"),
        "ingram_modified": metadata_date("Ingram", "ModifiedDate"),
        "amazon_sellthrough_report": metadata_date("Amazon Sellthrough", "ReportDate"),
        "bn_report": metadata_date("Barnes & Noble", "ReportDate"),
        "run_date": run_date,
    }

    return {
        spec.excel_column: date_rule_values.get(spec.row5_date_rule)
        for spec in SUPERCHARGED_COLUMNS
        if spec.row5_date_rule is not None
    }


def create_supercharged_frontlist_workbook(
    combined_df: pd.DataFrame,
    metadata_df: pd.DataFrame,
    frontlist_path: Path,
    output_dir: Path | None = None,
) -> Path:
    output_path = _resolve_supercharged_output_path(frontlist_path, output_dir)
    copy2(frontlist_path, output_path)

    workbook = load_workbook(output_path)
    if "Summary" not in workbook.sheetnames:
        raise ValueError(f"Workbook is missing the required 'Summary' sheet: {output_path}")

    worksheet = workbook["Summary"]
    insert_col_idx = column_index_from_string(SUMMARY_INSERT_AT_COLUMN)
    inserted_width = len(SUPERCHARGED_COLUMNS)
    isbn_col_idx = _find_column_by_header(worksheet, "ISBN-13")
    shifted_merged_ranges = _get_overlapping_merged_ranges(
        worksheet,
        min_col=insert_col_idx,
        max_col=worksheet.max_column,
        min_row=SUMMARY_HEADER_TOP_ROW,
        max_row=6,
    )

    _clear_overlapping_merged_ranges(
        worksheet,
        min_col=insert_col_idx,
        max_col=worksheet.max_column,
        min_row=SUMMARY_HEADER_TOP_ROW,
        max_row=6,
    )
    worksheet.delete_cols(insert_col_idx, 1)
    worksheet.insert_cols(insert_col_idx, inserted_width)
    _restore_shifted_merged_ranges(
        worksheet,
        shifted_merged_ranges,
        insert_col_idx=insert_col_idx + 1,
        inserted_width=inserted_width,
    )

    number_data_template = worksheet["M7"]
    date_data_template = worksheet["A7"]
    worksheet.row_dimensions[4].height = 14.45
    worksheet.row_dimensions[5].height = 14.45
    worksheet.row_dimensions[6].height = 60.75
    for spec in SUPERCHARGED_COLUMNS:
        worksheet.column_dimensions[spec.excel_column].width = spec.width

    _apply_supercharged_upper_fill(worksheet)

    for row_idx in (4, 5, 6):
        for spec in SUPERCHARGED_COLUMNS:
            cell = worksheet[f"{spec.excel_column}{row_idx}"]
            _apply_builtin_supercharged_header_style(
                cell,
                row_idx=row_idx,
                section=_find_supercharged_section(spec.excel_column),
            )

    for section in SUPERCHARGED_SECTIONS:
        worksheet[f"{section.start_column}4"].value = section.title

    row5_dates = _build_summary_row5_dates(metadata_df)
    for col_letter, value in row5_dates.items():
        worksheet[f"{col_letter}5"].value = value
        worksheet[f"{col_letter}5"].number_format = "mm/dd/yy;@"

    for spec in SUPERCHARGED_COLUMNS:
        worksheet[f"{spec.excel_column}6"].value = spec.header_label

    lookup_df = combined_df.copy()
    lookup_df["ISBN"] = lookup_df["ISBN"].astype("string")
    isbn_lookup = lookup_df.set_index("ISBN").to_dict("index")

    max_row = worksheet.max_row
    for row_idx in range(SUMMARY_FIRST_DATA_ROW, max_row + 1):
        isbn_value = _normalize_isbn_value(worksheet.cell(row=row_idx, column=isbn_col_idx).value)
        metrics = isbn_lookup.get(isbn_value)

        for spec in SUPERCHARGED_COLUMNS:
            target_cell = worksheet[f"{spec.excel_column}{row_idx}"]
            template_cell = date_data_template if spec.value_kind == "date" else number_data_template
            _copy_cell_style(target_cell, template_cell)
            if spec.value_kind == "date":
                # Use the same neutral border treatment as the numeric data cells.
                target_cell.border = copy(number_data_template.border)

            if not metrics or spec.source_column is None:
                target_cell.value = None
                continue

            value = metrics.get(spec.source_column)
            if spec.value_kind == "date":
                target_cell.number_format = "mm/dd/yy;@"
                if pd.isna(value):
                    target_cell.value = "-"
                else:
                    target_cell.value = pd.to_datetime(value).to_pydatetime()
            else:
                target_cell.value = None if pd.isna(value) else value

    _sort_summary_rows_by_column_desc(
        worksheet,
        sort_column_letter="AF",
        first_data_row=SUMMARY_FIRST_DATA_ROW,
    )

    workbook.save(output_path)
    return output_path


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Frontlist main data and SuperCharged workbook.")
    parser.add_argument("--frontlist-path", type=Path, help="Use a specific Frontlist Tracking workbook.")
    parser.add_argument(
        "--frontlist-output-dir",
        type=Path,
        help="Override where the FrontlistMain workbook is saved.",
    )
    parser.add_argument(
        "--supercharged-output-dir",
        type=Path,
        help="Override where the _SuperCharged workbook is saved.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    df, metadata_df, frontlist_path = build_frontlist_main(args.frontlist_path)
    output_path = save_frontlist_main_output(df, metadata_df, output_dir=args.frontlist_output_dir)
    supercharged_path = create_supercharged_frontlist_workbook(
        df,
        metadata_df,
        frontlist_path,
        output_dir=args.supercharged_output_dir,
    )

    print(f"Loaded Frontlist source: {frontlist_path}")
    print(f"Rows in combined output: {len(df)}")
    print(df.head(20).to_string(index=False))
    print("\nSource date summary:")
    print(metadata_df.to_string(index=False))
    print(f"\nSaved output to: {output_path}")
    print(f"Saved Frontlist SuperCharged workbook to: {supercharged_path}")


if __name__ == "__main__":
    main()
